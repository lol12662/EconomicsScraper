#!/usr/bin/env python3
"""Desktop GUI for the WSJ Treasury / Barchart Futures scraper."""

from __future__ import annotations

import sys
import threading
import traceback
from pathlib import Path
from tkinter import (
    BOTH, END, LEFT, RIGHT, X, Y, W,
    BooleanVar, Button, Entry, Frame, Label, StringVar, Text, Tk,
    filedialog, messagebox,
)
from tkinter import ttk

# ── Path resolution ────────────────────────────────────────────────────────────
if getattr(sys, "frozen", False):
    HERE = Path(sys.executable).resolve().parent
    _bundle = Path(getattr(sys, "_MEIPASS", HERE))
    for _p in (str(HERE), str(_bundle)):
        if _p not in sys.path:
            sys.path.insert(0, _p)
else:
    HERE = Path(__file__).resolve().parent
    if str(HERE) not in sys.path:
        sys.path.insert(0, str(HERE))

try:
    import wsj_treasury_scraper_fixed as scraper
except Exception as exc:
    scraper = None
    _import_error = exc
else:
    _import_error = None

# matplotlib — optional, chart tab is disabled if not installed
try:
    import matplotlib
    matplotlib.use("TkAgg")
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
    _MPL = True
except ImportError:
    _MPL = False

def _ensure_playwright_chromium(log_fn=None) -> bool:
    """
    If running as a frozen app and Playwright's Chromium isn't installed yet,
    install it automatically into the user's home directory.
    Returns True if Chromium is available, False if install failed.
    """
    import subprocess
    try:
        from playwright.sync_api import sync_playwright
        # Quick check: can we actually find a browser?
        with sync_playwright() as p:
            p.chromium.executable_path  # raises if not found
        return True
    except Exception:
        pass

    if log_fn:
        log_fn("First launch: installing Chromium browser (~150 MB)…")
        log_fn("This only happens once. Please wait…")

    try:
        result = subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            capture_output=True, text=True, timeout=300,
        )
        if result.returncode == 0:
            if log_fn: log_fn("✓ Chromium installed successfully.")
            return True
        else:
            if log_fn: log_fn(f"Chromium install failed:\n{result.stderr}")
            return False
    except Exception as e:
        if log_fn: log_fn(f"Chromium install error: {e}")
        return False


WSJ_URL      = "https://www.wsj.com/market-data/bonds/treasuries#treasuryB"
BARCHART_URL = "https://www.barchart.com/futures/financials?viewName=main"
SOURCE_WSJ      = "WSJ Treasuries"
SOURCE_BARCHART = "Barchart Futures"
SOURCE_MARKET   = "Market Data (Yahoo/FRED)"
SOURCES = [SOURCE_WSJ, SOURCE_BARCHART, SOURCE_MARKET]

# Columns to skip as chart series (used as X axis or non-numeric)
_SKIP_COLS = {"Maturity", "Article Date", "Last Payment Date", "Symbol",
              "Contract Name", "Time"}


class TreasuryApp(Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Treasury / Futures Scraper")
        self.geometry("1280x860")
        self.minsize(1000, 700)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self.source_var    = StringVar(value=SOURCE_WSJ)
        self.url_var       = StringVar(value=WSJ_URL)
        self.output_var    = StringVar(value=str(HERE / "wsj_treasuries.xlsx"))
        self.reference_var = StringVar(value="")
        self.delta_var          = StringVar(value="0.01")   # Delta: yield shift (decimal fraction)
        self._last_reference_date = ""    # stored after each WSJ scrape for chart title
        self.status_var    = StringVar(value="Ready.")

        self._worker: threading.Thread | None = None
        self._current_df = None
        self._current_bc_df = None
        self._current_mkt_df = None
        self._chart_vars: dict[str, BooleanVar] = {}      # WSJ column → checkbox var
        self._chart_bc_vars: dict[str, BooleanVar] = {}   # Barchart series → checkbox var
        self._chart_mkt_vars: dict[str, BooleanVar] = {}  # Market data series → checkbox var

        # Persist API key across sessions via a simple file
        self._api_key_file = HERE / ".fred_api_key"
        self._saved_api_key = ""
        try:
            if self._api_key_file.exists():
                self._saved_api_key = self._api_key_file.read_text().strip()
        except Exception:
            pass

        self._build_ui()
        if _import_error is not None:
            self._log(f"Import warning: {_import_error}")
            self._set_status("The scraper module could not be imported.")

    # ── UI ─────────────────────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        outer = Frame(self, padx=12, pady=12)
        outer.pack(fill=BOTH, expand=True)

        # ── Controls ───────────────────────────────────────────────────────────
        top = Frame(outer)
        top.pack(fill=X)

        Label(top, text="Data source").grid(row=0, column=0, sticky="w", pady=(0, 4))
        src_frame = Frame(top)
        src_frame.grid(row=0, column=1, sticky="w", padx=(8, 8), pady=(0, 4))
        for src in SOURCES:
            ttk.Radiobutton(src_frame, text=src, variable=self.source_var,
                            value=src, command=self._on_source_changed).pack(side=LEFT, padx=(0, 16))

        # ── WSJ / Barchart fields ──────────────────────────────────────────────
        self._ref_label = Label(top, text="Reference date (YYYY-MM-DD)")
        self._ref_label.grid(row=1, column=0, sticky="w", pady=(0, 4))
        self._ref_entry = Entry(top, textvariable=self.reference_var)
        self._ref_entry.grid(row=1, column=1, sticky="we", padx=(8, 8), pady=(0, 4))

        self._delta_label = Label(top, text="Delta (yield shift, e.g. 0.01 = 1%)")
        self._delta_label.grid(row=2, column=0, sticky="w", pady=(0, 4))
        self._delta_entry = Entry(top, textvariable=self.delta_var, width=12)
        self._delta_entry.grid(row=2, column=1, sticky="w", padx=(8, 8), pady=(0, 4))

        self._url_label = Label(top, text="URL")
        self._url_label.grid(row=3, column=0, sticky="w", pady=(0, 4))
        self._url_entry = Entry(top, textvariable=self.url_var)
        self._url_entry.grid(row=3, column=1, sticky="we", padx=(8, 8), pady=(0, 4))

        self._out_label = Label(top, text="Output file")
        self._out_label.grid(row=4, column=0, sticky="w", pady=(0, 4))
        self._out_row = Frame(top)
        self._out_row.grid(row=4, column=1, sticky="we", padx=(8, 8), pady=(0, 4))
        Entry(self._out_row, textvariable=self.output_var).pack(side=LEFT, fill=X, expand=True)
        Button(self._out_row, text="Browse…", command=self._browse_output).pack(side=RIGHT, padx=(8, 0))

        # ── Market Data fields (hidden by default) ─────────────────────────────
        self._mkt_fred_label = Label(top, text="FRED API Key")
        self._mkt_fred_label.grid(row=1, column=0, sticky="w", pady=(0, 4))
        self._mkt_fred_row = Frame(top)
        self._mkt_fred_row.grid(row=1, column=1, sticky="we", padx=(8, 8), pady=(0, 4))
        self._fred_key_var = StringVar(value=self._saved_api_key)
        Entry(self._mkt_fred_row, textvariable=self._fred_key_var, show="*").pack(side=LEFT, fill=X, expand=True)
        Button(self._mkt_fred_row, text="Save", command=self._save_api_key, width=6).pack(side=LEFT, padx=(4, 0))
        Label(self._mkt_fred_row, text="(fred.stlouisfed.org)", font=("", 7),
              foreground="gray").pack(side=LEFT, padx=(6, 0))

        self._mkt_tickers_label = Label(top, text="Yahoo Tickers")
        self._mkt_tickers_label.grid(row=2, column=0, sticky="w", pady=(0, 4))
        self._tickers_var = StringVar(value="HYG, SPY, ^TNX")
        self._mkt_tickers_entry = Entry(top, textvariable=self._tickers_var)
        self._mkt_tickers_entry.grid(row=2, column=1, sticky="we", padx=(8, 8), pady=(0, 4))

        self._mkt_fred_series_label = Label(top, text="FRED Series")
        self._mkt_fred_series_label.grid(row=3, column=0, sticky="w", pady=(0, 4))
        self._fred_series_var = StringVar(value="DGS1, DTB3, DGS10, DBAA")
        self._mkt_fred_series_entry = Entry(top, textvariable=self._fred_series_var)
        self._mkt_fred_series_entry.grid(row=3, column=1, sticky="we", padx=(8, 8), pady=(0, 4))

        self._mkt_dates_label = Label(top, text="Start / End date")
        self._mkt_dates_label.grid(row=4, column=0, sticky="w", pady=(0, 4))
        self._mkt_dates_row = Frame(top)
        self._mkt_dates_row.grid(row=4, column=1, sticky="we", padx=(8, 8), pady=(0, 4))
        self._mkt_start_var = StringVar(value="2025-01-02")
        self._mkt_end_var   = StringVar(value="2025-06-30")
        Entry(self._mkt_dates_row, textvariable=self._mkt_start_var, width=14).pack(side=LEFT)
        Label(self._mkt_dates_row, text=" → ").pack(side=LEFT)
        Entry(self._mkt_dates_row, textvariable=self._mkt_end_var, width=14).pack(side=LEFT)

        self._mkt_csv_label = Label(top, text="Output CSV")
        self._mkt_csv_label.grid(row=5, column=0, sticky="w", pady=(0, 4))
        self._mkt_csv_row = Frame(top)
        self._mkt_csv_row.grid(row=5, column=1, sticky="we", padx=(8, 8), pady=(0, 4))
        self._mkt_output_var = StringVar(value=str(HERE / "HighYieldData.csv"))
        Entry(self._mkt_csv_row, textvariable=self._mkt_output_var).pack(side=LEFT, fill=X, expand=True)
        Button(self._mkt_csv_row, text="Browse…", command=self._browse_mkt_output).pack(side=RIGHT, padx=(4, 0))

        # ── Action buttons ─────────────────────────────────────────────────────
        btn_row = Frame(top)
        btn_row.grid(row=7, column=1, sticky="w", padx=(8, 8), pady=(6, 2))
        self._run_btn = Button(btn_row, text="Run scrape", command=self._run_clicked, width=16)
        self._run_btn.pack(side=LEFT)
        Button(btn_row, text="Open output folder", command=self._open_output_folder, width=18).pack(side=LEFT, padx=(8, 0))

        log_btn_row = Frame(top)
        log_btn_row.grid(row=8, column=1, sticky="w", padx=(8, 8), pady=(2, 8))
        Button(log_btn_row, text="Save log…", command=self._save_log, width=14).pack(side=LEFT)
        Button(log_btn_row, text="Clear log",  command=self._clear_log, width=12).pack(side=LEFT, padx=(8, 0))

        top.columnconfigure(1, weight=1)

        # ── Middle: notebook + log ─────────────────────────────────────────────
        mid = Frame(outer)
        mid.pack(fill=BOTH, expand=True, pady=(8, 8))

        # Notebook (Preview + Chart tabs)
        self.notebook = ttk.Notebook(mid)
        self.notebook.pack(side=LEFT, fill=BOTH, expand=True)

        # ── Tab 1: Preview ─────────────────────────────────────────────────────
        preview_tab = Frame(self.notebook)
        self.notebook.add(preview_tab, text="Preview")

        yscroll = ttk.Scrollbar(preview_tab, orient="vertical")
        xscroll = ttk.Scrollbar(preview_tab, orient="horizontal")
        yscroll.pack(side=RIGHT, fill=Y)
        xscroll.pack(side="bottom", fill=X)

        self.tree = ttk.Treeview(preview_tab, show="headings", height=18,
                                 yscrollcommand=yscroll.set,
                                 xscrollcommand=xscroll.set)
        yscroll.config(command=self.tree.yview)
        xscroll.config(command=self.tree.xview)
        self.tree.pack(fill=BOTH, expand=True)

        # ── Tab 2: Chart (WSJ) ─────────────────────────────────────────────────
        chart_tab = Frame(self.notebook)
        self.notebook.add(chart_tab, text="Chart (WSJ)")

        if _MPL:
            # Left panel: series checkboxes + controls
            ctrl_frame = Frame(chart_tab, width=190)
            ctrl_frame.pack(side=LEFT, fill=Y, padx=(8, 4), pady=8)
            ctrl_frame.pack_propagate(False)

            Label(ctrl_frame, text="Series", font=("", 10, "bold")).pack(anchor=W, pady=(0, 2))

            # Select All / Deselect All buttons
            sel_row = Frame(ctrl_frame)
            sel_row.pack(anchor=W, fill=X, pady=(0, 6))
            Button(sel_row, text="Select All",   command=self._wsj_select_all,   width=10).pack(side=LEFT)
            Button(sel_row, text="Deselect All", command=self._wsj_deselect_all, width=10).pack(side=LEFT, padx=(4, 0))

            self._checkbox_frame = Frame(ctrl_frame)
            self._checkbox_frame.pack(fill=X, anchor=W)

            Label(ctrl_frame, text="X axis", font=("", 10, "bold")).pack(anchor=W, pady=(10, 4))
            self._x_var = StringVar(value="Maturity")
            self._x_menu = ttk.Combobox(ctrl_frame, textvariable=self._x_var,
                                        state="readonly", width=18)
            self._x_menu.pack(anchor=W)
            self._x_menu.bind("<<ComboboxSelected>>", lambda _: self._draw_chart())

            Button(ctrl_frame, text="Update chart", command=self._draw_chart,
                   width=16).pack(anchor=W, pady=(12, 0))

            # Right: matplotlib canvas
            canvas_frame = Frame(chart_tab)
            canvas_frame.pack(side=LEFT, fill=BOTH, expand=True, pady=8, padx=(0, 8))

            self._fig = Figure(figsize=(7, 4), dpi=96, tight_layout=True)
            self._ax  = self._fig.add_subplot(111)
            self._canvas = FigureCanvasTkAgg(self._fig, master=canvas_frame)
            self._canvas.get_tk_widget().pack(fill=BOTH, expand=True)
            NavigationToolbar2Tk(self._canvas, canvas_frame).pack(fill=X)

            self._ax.set_title("Run a WSJ scrape to populate the chart")
            self._ax.set_xlabel("Maturity")
            self._canvas.draw()
        else:
            Label(chart_tab,
                  text="matplotlib is not installed.\nRun:  pip install matplotlib",
                  justify="center", pady=40).pack(expand=True)

        # ── Tab 3: Chart (Barchart) ─────────────────────────────────────────────
        bc_chart_tab = Frame(self.notebook)
        self.notebook.add(bc_chart_tab, text="Chart (Barchart)")

        if _MPL:
            # Left panel
            bc_ctrl = Frame(bc_chart_tab, width=190)
            bc_ctrl.pack(side=LEFT, fill=Y, padx=(8, 4), pady=8)
            bc_ctrl.pack_propagate(False)

            Label(bc_ctrl, text="Series", font=("", 10, "bold")).pack(anchor=W, pady=(0, 2))

            bc_sel_row = Frame(bc_ctrl)
            bc_sel_row.pack(anchor=W, fill=X, pady=(0, 6))
            Button(bc_sel_row, text="Select All",   command=self._bc_select_all,   width=10).pack(side=LEFT)
            Button(bc_sel_row, text="Deselect All", command=self._bc_deselect_all, width=10).pack(side=LEFT, padx=(4, 0))

            self._bc_checkbox_frame = Frame(bc_ctrl)
            self._bc_checkbox_frame.pack(fill=X, anchor=W)

            Button(bc_ctrl, text="Update chart", command=self._draw_bc_chart,
                   width=16).pack(anchor=W, pady=(12, 0))

            # Right: matplotlib canvas
            bc_canvas_frame = Frame(bc_chart_tab)
            bc_canvas_frame.pack(side=LEFT, fill=BOTH, expand=True, pady=8, padx=(0, 8))

            self._bc_fig = Figure(figsize=(7, 4), dpi=96, tight_layout=True)
            self._bc_ax  = self._bc_fig.add_subplot(111)
            self._bc_canvas = FigureCanvasTkAgg(self._bc_fig, master=bc_canvas_frame)
            self._bc_canvas.get_tk_widget().pack(fill=BOTH, expand=True)
            NavigationToolbar2Tk(self._bc_canvas, bc_canvas_frame).pack(fill=X)

            self._bc_ax.set_title("Run a Barchart scrape to populate the chart")
            self._bc_canvas.draw()
        else:
            Label(bc_chart_tab,
                  text="matplotlib is not installed.\nRun:  pip install matplotlib",
                  justify="center", pady=40).pack(expand=True)

        # ── Tab 4: Chart (Market Data) ─────────────────────────────────────────
        mkt_chart_tab = Frame(self.notebook)
        self.notebook.add(mkt_chart_tab, text="Chart (Market Data)")

        if _MPL:
            mkt_chart_top = Frame(mkt_chart_tab)
            mkt_chart_top.pack(fill=X, padx=8, pady=(6, 0))
            Label(mkt_chart_top, text="Series  ").pack(side=LEFT)
            Button(mkt_chart_top, text="Select All",   command=self._mkt_select_all,   width=10).pack(side=LEFT)
            Button(mkt_chart_top, text="Deselect All", command=self._mkt_deselect_all, width=10).pack(side=LEFT, padx=(4, 0))
            self._mkt_checkbox_frame = Frame(mkt_chart_top)
            self._mkt_checkbox_frame.pack(side=LEFT, padx=(12, 0))

            self._mkt_fig = Figure(figsize=(7, 4), dpi=96, tight_layout=True)
            self._mkt_ax  = self._mkt_fig.add_subplot(111)
            self._mkt_canvas = FigureCanvasTkAgg(self._mkt_fig, master=mkt_chart_tab)
            self._mkt_canvas.get_tk_widget().pack(fill=BOTH, expand=True, pady=(6, 0))
            NavigationToolbar2Tk(self._mkt_canvas, mkt_chart_tab).pack(fill=X)
            self._mkt_ax.set_title("Select Market Data source and click Fetch Data")
            self._mkt_canvas.draw()
        else:
            Label(mkt_chart_tab, text="Install matplotlib to see chart.\npip install matplotlib",
                  justify="center").pack(expand=True)

        # ── Tab 5: FRED Data Preview ───────────────────────────────────────────
        fred_preview_tab = Frame(self.notebook)
        self.notebook.add(fred_preview_tab, text="FRED Preview")

        # Header row with column filter
        fred_hdr = Frame(fred_preview_tab)
        fred_hdr.pack(fill=X, padx=8, pady=(6, 4))
        Label(fred_hdr, text="Showing fetched FRED + Yahoo data  ").pack(side=LEFT)
        Label(fred_hdr, text="Filter columns:").pack(side=LEFT)
        self._fred_filter_var = StringVar(value="")
        self._fred_filter_var.trace_add("write", lambda *_: self._apply_fred_filter())
        Entry(fred_hdr, textvariable=self._fred_filter_var, width=20).pack(side=LEFT, padx=(4, 0))
        Button(fred_hdr, text="Clear", command=lambda: self._fred_filter_var.set(""),
               width=6).pack(side=LEFT, padx=(4, 0))
        self._fred_row_label = Label(fred_hdr, text="", foreground="gray")
        self._fred_row_label.pack(side=RIGHT)

        # Treeview with both scrollbars
        fred_tree_frame = Frame(fred_preview_tab)
        fred_tree_frame.pack(fill=BOTH, expand=True, padx=8, pady=(0, 8))

        fred_ysc = ttk.Scrollbar(fred_tree_frame, orient="vertical")
        fred_xsc = ttk.Scrollbar(fred_tree_frame, orient="horizontal")
        fred_ysc.pack(side=RIGHT, fill=Y)
        fred_xsc.pack(side="bottom", fill=X)

        self._fred_tree = ttk.Treeview(fred_tree_frame, show="headings",
                                        yscrollcommand=fred_ysc.set,
                                        xscrollcommand=fred_xsc.set)
        fred_ysc.config(command=self._fred_tree.yview)
        fred_xsc.config(command=self._fred_tree.xview)
        self._fred_tree.pack(fill=BOTH, expand=True)

        # ── Tab 6: Regression (Bond Returns) ────────────────────────────────────
        regress_tab = Frame(self.notebook)
        self.notebook.add(regress_tab, text="Regression")

        regress_top = Frame(regress_tab)
        regress_top.pack(fill=X, padx=8, pady=(8, 4))
        Label(regress_top, text="Runs OLS regressions on the fetched Market Data:",
              font=("", 9)).pack(side=LEFT)
        Button(regress_top, text="Run Regression", command=self._run_regression_clicked,
               width=16).pack(side=RIGHT)

        Label(regress_tab,
              text="Model 1:  RHYG ~ RSP        Model 2:  RHYG ~ RSP + CHI",
              font=("", 9, "italic"), foreground="gray").pack(anchor=W, padx=8, pady=(0, 6))

        regress_scroll = ttk.Scrollbar(regress_tab, orient="vertical")
        regress_scroll.pack(side=RIGHT, fill=Y)
        self._regress_text = Text(regress_tab, wrap="none", font=("Courier New", 9),
                                  yscrollcommand=regress_scroll.set)
        regress_scroll.config(command=self._regress_text.yview)
        self._regress_text.pack(fill=BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        self._regress_text.insert(END, "Fetch Market Data first, then click 'Run Regression'.\n")
        self._regress_text.config(state="disabled")

        # ── Log panel ──────────────────────────────────────────────────────────
        log_box = Frame(mid, width=380)
        log_box.pack(side=RIGHT, fill=Y, padx=(8, 0))
        log_box.pack_propagate(False)
        Label(log_box, text="Log").pack(anchor="w")
        log_scroll = ttk.Scrollbar(log_box, orient="vertical")
        self.log = Text(log_box, wrap="word", yscrollcommand=log_scroll.set)
        log_scroll.config(command=self.log.yview)
        log_scroll.pack(side=RIGHT, fill=Y)
        self.log.pack(fill=BOTH, expand=True)

        # ── Status bar ─────────────────────────────────────────────────────────
        bottom = Frame(outer)
        bottom.pack(fill=X)
        ttk.Separator(bottom, orient="horizontal").pack(fill=X, pady=(0, 6))
        Label(bottom, textvariable=self.status_var, anchor="w").pack(fill=X)

        self._log("Ready. Choose a data source, set an output file, and press Run scrape.")
        self._on_source_changed()

    # ── Helpers ────────────────────────────────────────────────────────────────

    def _on_close(self) -> None:
        if self._worker and self._worker.is_alive():
            if not messagebox.askyesno("Scrape running", "A scrape is still running. Quit anyway?"):
                return
        self.destroy()

    def _on_source_changed(self) -> None:
        src = self.source_var.get()
        wsj_bc = [self._url_label, self._url_entry, self._out_label, self._out_row]
        mkt = [
            self._mkt_fred_label, self._mkt_fred_row,
            self._mkt_tickers_label, self._mkt_tickers_entry,
            self._mkt_fred_series_label, self._mkt_fred_series_entry,
            self._mkt_dates_label, self._mkt_dates_row,
            self._mkt_csv_label, self._mkt_csv_row,
        ]
        if src == SOURCE_MARKET:
            for w in wsj_bc: w.grid_remove()
            self._ref_label.grid_remove()
            self._ref_entry.grid_remove()
            self._delta_label.grid_remove()
            self._delta_entry.grid_remove()
            for w in mkt: w.grid()
            self._run_btn.config(text="Fetch Data")
        else:
            for w in mkt: w.grid_remove()
            for w in wsj_bc: w.grid()
            self._run_btn.config(text="Run scrape")
            if src == SOURCE_WSJ:
                self.url_var.set(WSJ_URL)
                self.output_var.set(str(HERE / "wsj_treasuries.xlsx"))
                self._ref_label.grid()
                self._ref_entry.grid()
                self._delta_label.grid()
                self._delta_entry.grid()
            else:
                self.url_var.set(BARCHART_URL)
                self.output_var.set(str(HERE / "barchart_futures.xlsx"))
                self._ref_label.grid_remove()
                self._ref_entry.grid_remove()
                self._delta_label.grid_remove()
                self._delta_entry.grid_remove()


    def _set_status(self, text: str) -> None:
        self.status_var.set(text)

    def _log(self, text: str) -> None:
        self.log.insert(END, text.rstrip() + "\n")
        self.log.see(END)

    def _clear_log(self) -> None:
        self.log.delete("1.0", END)
        self._log("Log cleared.")

    def _save_log(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Save log", defaultextension=".txt",
            filetypes=[("Text file", "*.txt"), ("All files", "*.*")],
            initialfile="scraper_log.txt", initialdir=str(HERE),
        )
        if path:
            try:
                Path(path).write_text(self.log.get("1.0", END), encoding="utf-8")
                self._log(f"Log saved to {path}")
            except Exception as exc:
                messagebox.showerror("Save failed", str(exc))

    def _browse_output(self) -> None:
        filename = filedialog.asksaveasfilename(
            title="Choose output Excel file", defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            initialfile=Path(self.output_var.get()).name or "output.xlsx",
            initialdir=str(HERE),
        )
        if filename:
            self.output_var.set(filename)

    def _open_output_folder(self) -> None:
        path   = Path(self.output_var.get()).expanduser()
        folder = path.parent if path.suffix else path
        try:
            if sys.platform.startswith("win"):
                import os; os.startfile(str(folder))
            elif sys.platform == "darwin":
                import subprocess; subprocess.run(["open",     str(folder)], check=False)
            else:
                import subprocess; subprocess.run(["xdg-open", str(folder)], check=False)
        except Exception as exc:
            messagebox.showerror("Open folder failed", str(exc))

    def _run_clicked(self) -> None:
        if self._worker and self._worker.is_alive():
            messagebox.showinfo("Busy", "A scrape is already running.")
            return
        if scraper is None:
            messagebox.showerror("Import error",
                f"Could not import wsj_treasury_scraper: {_import_error}")
            return
        src = self.source_var.get()
        self._set_status(f"Running ({src})…")
        self._log(f"─── Starting {src} ───")
        if src == SOURCE_MARKET:
            target = self._run_mkt
        elif src == SOURCE_BARCHART:
            target = self._run_barchart
        else:
            target = self._run_wsj
        self._worker = threading.Thread(target=target, daemon=True)
        self._worker.start()

    # ── WSJ scrape ─────────────────────────────────────────────────────────────

    def _run_wsj(self) -> None:
        try:
            url    = self.url_var.get().strip()
            output = Path(self.output_var.get().strip()).expanduser()
            ref_raw = self.reference_var.get().strip()
            reference_date = scraper.get_reference_date(url, ref_raw or None)

            # Parse Delta (default 0.01 = 1% if blank or invalid)
            try:
                delta = float(self.delta_var.get().strip())
                if delta <= 0:
                    raise ValueError("Delta must be positive")
            except ValueError:
                delta = 0.01
                self.after(0, lambda: self._log("Invalid Delta — using default 0.01 (1%)"))

            delta_pct = delta * 100
            ref_iso = reference_date.isoformat()
            self.after(0, lambda r=ref_iso: setattr(self, "_last_reference_date", r))
            self.after(0, lambda r=ref_iso: self._log(f"Reference date: {r}"))
            self.after(0, lambda dp=delta_pct: self._log(f"Delta: {delta} ({dp:.4g}% yield shift)"))
            self.after(0, lambda: self._log("Scraping Treasury table…"))

            df = scraper.scrape_treasuries(url)
            df = scraper.add_payment_columns(df, reference_date, delta=delta)

            output.parent.mkdir(parents=True, exist_ok=True)
            self._write_wsj_excel(df, output, delta, delta_pct, reference_date=reference_date)
            self.after(0, lambda: self._log("  Regression sheet written to Sheet 2."))

            n = len(df)
            self.after(0, lambda: self._update_preview(df, is_wsj=True))
            self.after(0, lambda: self._log(f"✓ Saved {n} rows → {output.resolve()}"))
            self.after(0, lambda: self._set_status(f"Done — {n} rows saved to {output.name}"))
        except Exception as exc:
            tb = traceback.format_exc()
            self.after(0, lambda: self._log(tb))
            self.after(0, lambda: self._set_status("Failed — see log for details."))

    def _write_wsj_excel(self, df, output: Path, delta: float, delta_pct: float,
                          reference_date=None) -> None:
        """Write Sheet 1 (data + formula legend) and Sheet 2 (regressions) to Excel."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import date as _date

        # Row 1: date header, Row 2: blank, Rows 3-16: formula legend, Row 17+: data
        formula_rows = 16
        sheet_name   = "Treasuries"

        # Determine display date
        if reference_date is not None:
            display_date = reference_date.strftime("%B %d, %Y")
            date_label   = f"Reference Date: {display_date}"
        else:
            display_date = _date.today().strftime("%B %d, %Y")
            date_label   = f"Date: {display_date}"

        with scraper.pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=formula_rows)
            ws = writer.book[sheet_name]

            # ── Row 1: Date header ────────────────────────────────────────────
            ws["A1"] = date_label
            ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
            ws["A1"].fill      = PatternFill("solid", fgColor="1F4E79")
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 22

            # ── Rows 3–16: Formula legend ─────────────────────────────────────
            ws["A3"]  = "Treasury Formulas"
            ws["A3"].font = Font(bold=True)
            ws["A4"]  = "Coupon Payment";    ws["B4"]  = "Coupon/2*1000"
            ws["A5"]  = "PV0";               ws["B5"]  = "PV(Ask Yield/2, Number of Payments Until Maturity, Coupon Payment, 1000)"
            ws["A6"]  = "PV1";               ws["B6"]  = f"PV((Ask Yield+{delta_pct:.4g}%)/2, Number of Payments Until Maturity-2, Coupon Payment, 1000)"
            ws["A7"]  = "P0";                ws["B7"]  = "PV0 * (1+Asked Yield/2)^(Days Since Last Payment/182)"
            ws["A8"]  = "P1";                ws["B8"]  = f"PV1 * (1+(Asked Yield+{delta_pct:.4g}%)/2)^(Days Since Last Payment/182)"
            ws["A9"]  = "Simulated Return";  ws["B9"]  = "(P1- P0 + Coupon*10)/P0"
            ws["A10"] = "MACAULAY DURATION"; ws["B10"] = "DURATION(Current Date, Maturity Date, Coupon Rate, Ask Yield, 2)"
            ws["A11"] = "MODIFIED DURATION"; ws["B11"] = "MDURATION(Current Date, Maturity Date, Coupon Rate, Ask Yield, 2)"
            ws["A12"] = "PVUP";              ws["B12"] = f"PV(Ask Yield+{delta_pct:.4g}%/2, Number of Payments Until Maturity, Coupon Payment, 1000)"
            ws["A13"] = "PVDN";              ws["B13"] = f"PV(Ask Yield-{delta_pct:.4g}%/2, Number of Payments Until Maturity, Coupon Payment, 1000)"
            ws["A14"] = "PUP";               ws["B14"] = f"PVUP * (1+(Asked Yield+{delta_pct:.4g}%)/2)^(Days Since Last Payment/182)"
            ws["A15"] = "PDN";               ws["B15"] = f"PVDN * (1+(Asked Yield-{delta_pct:.4g}%)/2)^(Days Since Last Payment/182)"
            ws["A16"] = "EFDURATION";        ws["B16"] = f"(PDN- PUP)/(2*{delta}*P0)"

            # ── Sheet 2: Regressions ──────────────────────────────────────────
            try:
                reg_df = scraper.run_wsj_regressions(df)
                reg_df.to_excel(writer, index=False, sheet_name="Regressions")
                reg_ws = writer.book["Regressions"]

                header_fill = PatternFill("solid", fgColor="D9E1F2")
                bold_font   = Font(bold=True)
                for row in reg_ws.iter_rows():
                    if row[0].value and str(row[0].value).startswith("Ask Yield"):
                        for cell in row:
                            cell.font = bold_font
                            cell.fill = header_fill
                for col in reg_ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=8)
                    reg_ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
            except Exception as reg_exc:
                pass   # regression failure doesn't prevent saving the main data

            # ── Sheet 3: Regression Charts ────────────────────────────────────
            try:
                chart_png = scraper.build_regression_chart_image(df)
                if chart_png is not None:
                    from openpyxl.drawing.image import Image as XLImage
                    from openpyxl.styles import Font as XLFont
                    import io
                    chart_ws = writer.book.create_sheet("Regression Charts")
                    chart_ws["A1"] = "Ask Yield Regression Plots"
                    chart_ws["A1"].font = XLFont(bold=True, size=13)
                    img = XLImage(io.BytesIO(chart_png))
                    img.anchor = "A3"
                    chart_ws.add_image(img)
            except Exception:
                pass   # chart failure doesn't prevent saving

    # ── Barchart scrape ────────────────────────────────────────────────────────

    def _run_barchart(self) -> None:
        try:
            url    = self.url_var.get().strip()
            output = Path(self.output_var.get().strip()).expanduser()

            self.after(0, lambda: self._log("Scraping Barchart futures table…"))
            # Ensure Chromium is available (installs on first run if needed)
            _ensure_playwright_chromium(log_fn=lambda m: self.after(0, lambda m=m: self._log(m)))
            df = scraper.scrape_barchart_futures(
                url, log=lambda m: self.after(0, lambda m=m: self._log(m)),
            )

            output.parent.mkdir(parents=True, exist_ok=True)
            with scraper.pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Barchart Futures")

            n = len(df)
            self.after(0, lambda: self._update_preview(df, is_wsj=False))
            self.after(0, lambda: self._log(f"✓ Saved {n} rows → {output.resolve()}"))
            self.after(0, lambda: self._set_status(f"Done — {n} rows saved to {output.name}"))
        except Exception as exc:
            tb = traceback.format_exc()
            self.after(0, lambda: self._log(tb))
            self.after(0, lambda: self._set_status("Failed — see log for details."))

    # ── Preview & Chart ────────────────────────────────────────────────────────

    def _update_preview(self, df, is_wsj: bool = False) -> None:
        self._current_df = df
        cols = list(df.columns)

        # Update treeview
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = cols
        COL_WIDTH = 130
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=COL_WIDTH, minwidth=COL_WIDTH,
                             stretch=False, anchor="w")
        for _, row in df.head(100).iterrows():
            self.tree.insert("", END, values=[str(v) for v in row.tolist()])
        self._log(f"Preview updated ({min(len(df), 100)} of {len(df)} rows shown).")

        # Update chart controls
        if _MPL and is_wsj:
            self._rebuild_chart_controls(df)
            self._draw_chart()
        elif _MPL and not is_wsj:
            self._current_bc_df = df
            self._rebuild_bc_chart_controls(df)
            self._draw_bc_chart()

    def _rebuild_chart_controls(self, df) -> None:
        """Rebuild the series checkboxes and X-axis dropdown from the new dataframe."""
        # Clear old checkboxes
        for widget in self._checkbox_frame.winfo_children():
            widget.destroy()
        self._chart_vars.clear()

        # Numeric columns only (exclude known non-numeric / date cols)
        import pandas as pd
        numeric_cols = [
            c for c in df.columns
            if c not in _SKIP_COLS and pd.api.types.is_numeric_dtype(
                pd.to_numeric(df[c], errors="coerce")
            )
        ]

        for col in numeric_cols:
            var = BooleanVar(value=True)
            self._chart_vars[col] = var
            ttk.Checkbutton(self._checkbox_frame, text=col, variable=var,
                            command=self._draw_chart).pack(anchor=W)

        # X axis dropdown — date/text columns
        x_options = [c for c in df.columns if c not in numeric_cols]
        self._x_menu["values"] = x_options
        if "Maturity" in x_options:
            self._x_var.set("Maturity")
        elif x_options:
            self._x_var.set(x_options[0])

    # ── WSJ select / deselect all ──────────────────────────────────────────────
    def _wsj_select_all(self) -> None:
        for var in self._chart_vars.values():
            var.set(True)
        self._draw_chart()

    def _wsj_deselect_all(self) -> None:
        for var in self._chart_vars.values():
            var.set(False)
        self._draw_chart()

    # ── Barchart select / deselect all ─────────────────────────────────────────
    def _bc_select_all(self) -> None:
        for var in self._chart_bc_vars.values():
            var.set(True)
        self._draw_bc_chart()

    def _bc_deselect_all(self) -> None:
        for var in self._chart_bc_vars.values():
            var.set(False)
        self._draw_bc_chart()

    def _rebuild_bc_chart_controls(self, df) -> None:
        """Rebuild Barchart series checkboxes — Latest and Change only."""
        for widget in self._bc_checkbox_frame.winfo_children():
            widget.destroy()
        self._chart_bc_vars.clear()

        for col in ("Latest", "Change"):
            if col in df.columns:
                var = BooleanVar(value=True)
                self._chart_bc_vars[col] = var
                ttk.Checkbutton(self._bc_checkbox_frame, text=col, variable=var,
                                command=self._draw_bc_chart).pack(anchor=W)

    def _draw_bc_chart(self) -> None:
        """Draw a line chart of Latest / Change by Contract Name."""
        if not _MPL or self._current_bc_df is None:
            return

        import pandas as pd

        df = self._current_bc_df.copy()
        selected = [col for col, var in self._chart_bc_vars.items() if var.get()]

        self._bc_ax.clear()

        if not selected:
            self._bc_ax.set_title("No series selected")
            self._bc_canvas.draw()
            return

        # Use Contract Name as X labels, fall back to Symbol
        if "Contract Name" in df.columns:
            labels = df["Contract Name"].astype(str).tolist()
        elif "Symbol" in df.columns:
            labels = df["Symbol"].astype(str).tolist()
        else:
            labels = [str(i) for i in range(len(df))]

        x = range(len(labels))

        for col in selected:
            y_vals = pd.to_numeric(df[col], errors="coerce")
            self._bc_ax.plot(list(x), y_vals.tolist(), marker="o",
                             markersize=5, linewidth=1.5, label=col)

        self._bc_ax.set_xticks(list(x))
        self._bc_ax.set_xticklabels(labels, rotation=40, ha="right", fontsize=8)
        self._bc_ax.axhline(0, color="black", linewidth=0.6, linestyle="--")
        self._bc_ax.set_title("Barchart Futures — Latest & Change")
        self._bc_ax.set_ylabel("Price (decimal)")
        self._bc_ax.legend(fontsize=9)
        self._bc_ax.grid(True, linestyle="--", alpha=0.4)
        self._bc_fig.tight_layout()
        self._bc_canvas.draw()
        # Switch to barchart chart tab
        self.notebook.select(2)

    def _draw_chart(self) -> None:
        """Redraw the matplotlib chart based on current checkbox selections."""
        if not _MPL or self._current_df is None:
            return

        import pandas as pd
        import matplotlib.dates as mdates
        import matplotlib.ticker as mticker

        df = self._current_df.copy()
        x_col = self._x_var.get()
        selected = [col for col, var in self._chart_vars.items() if var.get()]

        self._ax.clear()

        if not selected:
            self._ax.set_title("No series selected")
            self._canvas.draw()
            return

        if x_col not in df.columns:
            self._ax.set_title(f"X column '{x_col}' not found")
            self._canvas.draw()
            return

        # Try to parse X as dates so we get a proper time axis
        x_as_dates = pd.to_datetime(df[x_col], errors="coerce")
        use_dates = x_as_dates.notna().sum() > len(df) * 0.5

        if use_dates:
            df["_x"] = x_as_dates
        else:
            df["_x"] = range(len(df))

        # Sort by X so the line runs left→right
        df = df.sort_values("_x").reset_index(drop=True)

        for col in selected:
            y_vals = pd.to_numeric(df[col], errors="coerce")
            self._ax.plot(df["_x"], y_vals, marker="o", markersize=3,
                         linewidth=1.2, label=col)

        self._ax.set_xlabel(x_col)
        ref = getattr(self, "_last_reference_date", "")
        try:
            delta_val = float(self.delta_var.get().strip())
        except ValueError:
            delta_val = 0.01
        title = "WSJ Treasury Data"
        if ref:
            title += f"  |  Reference Date: {ref}  |  Delta: {delta_val} ({delta_val*100:.4g}%)"
        self._ax.set_title(title, fontsize=9)
        self._ax.legend(fontsize=8, loc="best")
        self._ax.grid(True, linestyle="--", alpha=0.4)

        if use_dates:
            # Auto-format: pick a sensible locator based on date range
            date_range = (df["_x"].max() - df["_x"].min()).days
            if date_range > 365 * 5:
                self._ax.xaxis.set_major_locator(mdates.YearLocator(2))
                self._ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
            elif date_range > 365:
                self._ax.xaxis.set_major_locator(mdates.YearLocator())
                self._ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
            else:
                self._ax.xaxis.set_major_locator(mdates.MonthLocator())
                self._ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
        else:
            # Numeric index: show at most 20 evenly-spaced ticks
            n = len(df)
            step = max(1, n // 20)
            ticks = list(range(0, n, step))
            labels = [str(df[x_col].iloc[i]) for i in ticks]
            self._ax.set_xticks(ticks)
            self._ax.set_xticklabels(labels)

        self._fig.autofmt_xdate(rotation=45, ha="right")
        self._canvas.draw()
        # Switch to chart tab automatically
        self.notebook.select(1)


    # ── Market Data helpers ────────────────────────────────────────────────────

    def _save_api_key(self) -> None:
        key = self._fred_key_var.get().strip()
        try:
            self._api_key_file.write_text(key)
            self._saved_api_key = key
            self._log("FRED API key saved.")
        except Exception as e:
            messagebox.showerror("Save failed", str(e))

    def _browse_mkt_output(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Save CSV as",
            defaultextension=".csv",
            filetypes=[("CSV file", "*.csv"), ("All files", "*.*")],
            initialfile=Path(self._mkt_output_var.get()).name,
            initialdir=str(HERE),
        )
        if path:
            self._mkt_output_var.set(path)

    def _run_mkt(self) -> None:
        if not self._fred_key_var.get().strip():
            self.after(0, lambda: messagebox.showerror(
                "Missing API key",
                "Please enter your FRED API key.\nGet one free at fred.stlouisfed.org"))
            self.after(0, lambda: self._set_status("Ready."))
            return
        try:
            import pandas as pd
            try:
                import yfinance as yf
            except ImportError:
                self.after(0, lambda: self._log("Installing yfinance…"))
                import subprocess
                subprocess.run([sys.executable, "-m", "pip", "install", "yfinance"],
                               capture_output=True)
                import yfinance as yf

            try:
                from fredapi import Fred
            except ImportError:
                self.after(0, lambda: self._log("Installing fredapi…"))
                import subprocess
                subprocess.run([sys.executable, "-m", "pip", "install", "fredapi"],
                               capture_output=True)
                from fredapi import Fred

            api_key    = self._fred_key_var.get().strip()
            tickers    = [t.strip() for t in self._tickers_var.get().split(",") if t.strip()]
            fred_series= [s.strip() for s in self._fred_series_var.get().split(",") if s.strip()]
            start      = self._mkt_start_var.get().strip()
            end        = self._mkt_end_var.get().strip()
            output     = Path(self._mkt_output_var.get().strip()).expanduser()

            # Yahoo Finance
            self.after(0, lambda: self._log(f"Downloading Yahoo Finance: {tickers}…"))
            yahoo_data = yf.download(
                tickers, start=start, end=end,
                auto_adjust=False, progress=False,
            )
            # Handle single vs multiple tickers
            if len(tickers) == 1:
                yahoo_data = yahoo_data[["Adj Close"]].rename(columns={"Adj Close": tickers[0]})
            else:
                yahoo_data = yahoo_data["Adj Close"]
            yahoo_data = yahoo_data.dropna()
            n_yahoo = len(yahoo_data)
            self.after(0, lambda: self._log(f"  ✓ Yahoo: {n_yahoo} rows, {list(yahoo_data.columns)}"))

            # FRED
            self.after(0, lambda: self._log("Downloading FRED data…"))
            fred = Fred(api_key=api_key)
            fred_data = pd.DataFrame()
            for series in fred_series:
                try:
                    fred_data[series] = fred.get_series(series, start, end)
                    self.after(0, lambda s=series: self._log(f"  ✓ FRED: {s}"))
                except Exception as e:
                    self.after(0, lambda s=series, e=e: self._log(f"  ✗ FRED: {s} — {e}"))

            # Merge
            combined = yahoo_data.join(fred_data, how="inner").dropna()
            combined = combined.reset_index()
            combined.rename(columns={combined.columns[0]: "Time"}, inplace=True)
            combined["Time"] = pd.to_datetime(combined["Time"]).dt.strftime("%Y%m%d")

            # Save
            output.parent.mkdir(parents=True, exist_ok=True)
            combined.to_csv(output, index=False)
            n = len(combined)
            cols = list(combined.columns)

            self.after(0, lambda: self._current_mkt_df_set(combined))
            self.after(0, lambda: self._log(f"✓ Saved {n} rows to {output.resolve()}"))
            self.after(0, lambda: self._log(f"  Columns: {cols}"))
            self.after(0, lambda: self._set_status(f"Done — {n} rows saved to {output.name}"))

        except Exception as exc:
            tb = traceback.format_exc()
            self.after(0, lambda: self._log(tb))
            self.after(0, lambda: self._set_status("Failed — see log for details."))

    def _current_mkt_df_set(self, df) -> None:
        """Called on main thread after fetch completes."""
        self._current_mkt_df = df
        self._populate_fred_preview(df)
        if _MPL:
            self._rebuild_mkt_chart_controls(df)
            self._draw_mkt_chart()

    def _populate_fred_preview(self, df) -> None:
        """Fill the FRED Preview treeview with the full combined dataframe."""
        self._fred_preview_df = df.copy()
        self._apply_fred_filter()

    def _apply_fred_filter(self) -> None:
        """Re-render the FRED treeview, optionally filtering visible columns."""
        if not hasattr(self, "_fred_preview_df") or self._fred_preview_df is None:
            return
        df = self._fred_preview_df
        filt = self._fred_filter_var.get().strip().lower()

        # Determine which columns to show
        if filt:
            cols = [c for c in df.columns if filt in c.lower()]
            # Always keep Time if it exists
            if "Time" in df.columns and "Time" not in cols:
                cols = ["Time"] + cols
        else:
            cols = list(df.columns)

        if not cols:
            self._fred_row_label.config(text="No columns match filter")
            return

        view = df[cols]

        # Rebuild treeview columns
        self._fred_tree.delete(*self._fred_tree.get_children())
        self._fred_tree["columns"] = cols
        COL_W = 110
        for col in cols:
            self._fred_tree.heading(col, text=col,
                                    command=lambda c=col: self._fred_sort(c))
            self._fred_tree.column(col, width=max(COL_W, len(col) * 9),
                                   minwidth=70, stretch=False, anchor="center")

        # Insert rows (all rows — treeview scrolls)
        for _, row in view.iterrows():
            self._fred_tree.insert("", END, values=[str(v) for v in row.tolist()])

        n = len(view)
        self._fred_row_label.config(
            text=f"{n} rows × {len(cols)} cols")
        self.notebook.select(4)   # jump to FRED Preview tab

    def _fred_sort(self, col: str) -> None:
        """Click a column header to sort by that column."""
        if not hasattr(self, "_fred_preview_df") or self._fred_preview_df is None:
            return
        import pandas as pd
        df = self._fred_preview_df
        try:
            # Try numeric sort, fall back to string
            sorted_df = df.sort_values(
                col,
                key=lambda s: pd.to_numeric(s, errors="coerce").fillna(float("inf")),
            )
        except Exception:
            sorted_df = df.sort_values(col, key=lambda s: s.astype(str))
        self._fred_preview_df = sorted_df.reset_index(drop=True)
        self._apply_fred_filter()

    def _rebuild_mkt_chart_controls(self, df) -> None:
        import pandas as pd
        for widget in self._mkt_checkbox_frame.winfo_children():
            widget.destroy()
        self._chart_mkt_vars.clear()

        skip = {"Time"}
        for col in df.columns:
            if col in skip:
                continue
            var = BooleanVar(value=True)
            self._chart_mkt_vars[col] = var
            ttk.Checkbutton(self._mkt_checkbox_frame, text=col, variable=var,
                            command=self._draw_mkt_chart).pack(side=LEFT)

    def _mkt_select_all(self) -> None:
        for var in self._chart_mkt_vars.values():
            var.set(True)
        self._draw_mkt_chart()

    def _mkt_deselect_all(self) -> None:
        for var in self._chart_mkt_vars.values():
            var.set(False)
        self._draw_mkt_chart()

    def _draw_mkt_chart(self) -> None:
        if not _MPL or self._current_mkt_df is None:
            return

        import pandas as pd
        import matplotlib.dates as mdates

        df = self._current_mkt_df.copy()
        selected = [col for col, var in self._chart_mkt_vars.items() if var.get()]

        self._mkt_ax.clear()

        if not selected:
            self._mkt_ax.set_title("No series selected")
            self._mkt_canvas.draw()
            return

        df["_date"] = pd.to_datetime(df["Time"], format="%Y%m%d", errors="coerce")
        df = df.sort_values("_date")

        for col in selected:
            y = pd.to_numeric(df[col], errors="coerce")
            self._mkt_ax.plot(df["_date"], y, linewidth=1.2, label=col)

        self._mkt_ax.set_title("Market Data — Yahoo Finance & FRED")
        self._mkt_ax.set_xlabel("Date")
        self._mkt_ax.legend(fontsize=8, loc="best")
        self._mkt_ax.grid(True, linestyle="--", alpha=0.4)

        date_range = (df["_date"].max() - df["_date"].min()).days
        if date_range > 365 * 5:
            self._mkt_ax.xaxis.set_major_locator(mdates.YearLocator(2))
            self._mkt_ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
        elif date_range > 365:
            self._mkt_ax.xaxis.set_major_locator(mdates.YearLocator())
            self._mkt_ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
        else:
            self._mkt_ax.xaxis.set_major_locator(mdates.MonthLocator())
            self._mkt_ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))

        self._mkt_fig.autofmt_xdate(rotation=45, ha="right")
        self._mkt_canvas.draw()
        self.notebook.select(3)

    # ── Regression (Bond_Regress) ──────────────────────────────────────────────

    def _run_regression_clicked(self) -> None:
        if self._worker and self._worker.is_alive():
            messagebox.showinfo("Busy", "A task is already running.")
            return
        if self._current_mkt_df is None:
            messagebox.showinfo("No data", "Fetch Market Data first (HYG, SPY, ^TNX, etc.).")
            return
        self._set_status("Running regression…")
        self._log("─── Starting bond return regression ───")
        self._worker = threading.Thread(target=self._run_regression, daemon=True)
        self._worker.start()

    def _run_regression(self) -> None:
        try:
            import pandas as pd
            try:
                import statsmodels.formula.api as smf
            except ImportError:
                self.after(0, lambda: self._log("Installing statsmodels…"))
                import subprocess
                subprocess.run([sys.executable, "-m", "pip", "install", "statsmodels"],
                               capture_output=True)
                import statsmodels.formula.api as smf

            df = self._current_mkt_df.copy()

            # Map whatever Yahoo columns are present to the names Bond_Regress expects.
            # HYG / SPY / ^TNX are the defaults, but allow flexibility if the user
            # changed the ticker list.
            rename_map = {}
            for col in df.columns:
                if col == "^TNX":
                    rename_map[col] = "TNX"
            df = df.rename(columns=rename_map)

            required = {"HYG", "SPY", "TNX"}
            missing = required - set(df.columns)
            if missing:
                raise ValueError(
                    f"Regression requires columns {sorted(required)} but missing {sorted(missing)}.\n"
                    f"Make sure your Yahoo Tickers include HYG, SPY, and ^TNX before fetching."
                )

            df["Date"] = pd.to_datetime(df["Time"], format="%Y%m%d")
            df = df.sort_values("Date").reset_index(drop=True)

            df["RHYG"] = df["HYG"] / df["HYG"].shift(1) - 1
            df["RSP"]  = df["SPY"] / df["SPY"].shift(1) - 1
            df["CHI"]  = df["TNX"] - df["TNX"].shift(1)

            df_ret = df[["Date", "RHYG", "RSP", "CHI"]].dropna().reset_index(drop=True)

            self.after(0, lambda: self._log(f"Computed returns for {len(df_ret)} rows."))

            lines = []
            lines.append("First 7 rows of return data:")
            lines.append(df_ret.head(7).to_string(index=False))
            lines.append("")
            lines.append("Last 7 rows of return data:")
            lines.append(df_ret.tail(7).to_string(index=False))
            lines.append("")
            lines.append("=" * 78)
            lines.append("Model 1: RHYG ~ RSP")
            lines.append("=" * 78)
            model1 = smf.ols("RHYG ~ RSP", data=df_ret).fit()
            lines.append(str(model1.summary()))
            lines.append("")
            lines.append("=" * 78)
            lines.append("Model 2: RHYG ~ RSP + CHI")
            lines.append("=" * 78)
            model2 = smf.ols("RHYG ~ RSP + CHI", data=df_ret).fit()
            lines.append(str(model2.summary()))

            output_text = "\n".join(lines)
            self.after(0, lambda: self._show_regression_results(output_text))
            self.after(0, lambda: self._log("✓ Regression complete."))
            self.after(0, lambda: self._set_status("Regression complete — see Regression tab."))

        except Exception as exc:
            tb = traceback.format_exc()
            self.after(0, lambda: self._log(tb))
            self.after(0, lambda: self._set_status("Failed — see log for details."))

    def _show_regression_results(self, text: str) -> None:
        self._regress_text.config(state="normal")
        self._regress_text.delete("1.0", END)
        self._regress_text.insert(END, text)
        self._regress_text.config(state="disabled")
        self.notebook.select(4)


def _write_crash_log(text: str) -> None:
    for path in [Path.home() / "Desktop" / "WSJ_Treasury_CRASH.txt",
                 HERE / "WSJ_Treasury_CRASH.txt"]:
        try:
            path.write_text(text, encoding="utf-8")
        except Exception:
            pass


def main() -> int:
    import traceback as _tb
    startup_info = (
        f"Python: {sys.version}\n"
        f"Executable: {sys.executable}\n"
        f"Frozen: {getattr(sys, 'frozen', False)}\n"
        f"HERE: {HERE}\n"
        f"scraper loaded: {scraper is not None}\n"
        f"import error: {_import_error}\n"
        f"matplotlib: {_MPL}\n"
    )
    try:
        if scraper is None:
            _write_crash_log(f"Could not import scraper:\n{_import_error}\n\n{startup_info}")
            try:
                import tkinter as _tk; _root = _tk.Tk(); _root.withdraw()
                from tkinter import messagebox as _mb
                _mb.showerror("Import Error", f"Could not load scraper:\n{_import_error}")
                _root.destroy()
            except Exception:
                pass
            return 1
        app = TreasuryApp()
        app.mainloop()
        return 0
    except Exception as exc:
        _write_crash_log(f"CRASH:\n{_tb.format_exc()}\n\n{startup_info}")
        try:
            import tkinter as _tk; _root = _tk.Tk(); _root.withdraw()
            from tkinter import messagebox as _mb
            _mb.showerror("Crash", f"{exc}\n\nSee WSJ_Treasury_CRASH.txt on Desktop.")
            _root.destroy()
        except Exception:
            pass
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
