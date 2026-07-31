#!/usr/bin/env python3
"""
Scrape WSJ Treasury data and export:
- Maturity
- Coupon
- Asked Yield
- Article Date
- Last Payment Date
- Days Since Last Payment
- Payments Until Maturity
- Coupon Payment
- PV

Install:
    python -m pip install pandas openpyxl requests playwright
    python -m playwright install chromium

Run:
    python wsj_treasury_scraper.py --output treasuries.xlsx
    python wsj_treasury_scraper.py --output treasuries.xlsx --reference-date 2026-05-18
"""

from __future__ import annotations

import argparse
import calendar
import json
import os
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any, Iterable, List, Optional, Sequence

import pandas as pd
import requests

DEFAULT_URL = "https://www.wsj.com/market-data/bonds/treasuries#treasuryB"
BARCHART_URL = "https://www.barchart.com/futures/financials?viewName=main"
TARGET_COLUMNS = ["Maturity", "Coupon", "Asked Yield"]

# Barchart futures columns
BARCHART_COLUMNS = ["Symbol", "Contract Name", "Latest", "Change", "Volume", "Time"]


# ============================================================
# PLAYWRIGHT BROWSER PATH RESOLUTION
# ============================================================

def _configure_playwright_browsers() -> None:
    """Point Playwright at a browser directory that works in frozen EXEs.

    Priority order:
    1) a bundled `ms-playwright` folder next to the EXE / inside _MEIPASS
    2) the standard user install location on Windows
    3) the standard cache locations on other platforms
    """
    if os.environ.get("PLAYWRIGHT_BROWSERS_PATH"):
        return

    candidates = []

    frozen_root = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    candidates.append(frozen_root / "ms-playwright")
    candidates.append(Path(__file__).resolve().parent / "ms-playwright")

    localappdata = os.environ.get("LOCALAPPDATA")
    if localappdata:
        candidates.append(Path(localappdata) / "ms-playwright")

    if os.name == "nt":
        appdata = os.environ.get("APPDATA")
        if appdata:
            candidates.append(Path(appdata) / "ms-playwright")

    candidates.append(Path.home() / ".cache" / "ms-playwright")
    candidates.append(Path.home() / ".cache" / "ms-playwright" / "chromium")

    for candidate in candidates:
        try:
            if candidate.exists():
                os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(candidate)
                return
        except Exception:
            continue


_configure_playwright_browsers()


# ============================================================
# DATE HELPERS
# ============================================================

def _safe_date(year: int, month: int, day: int) -> date:
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, min(day, last_day))


def parse_any_date(value) -> date:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        raise ValueError("Date value is empty")

    if isinstance(value, date):
        return value

    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        raise ValueError(f"Could not parse date value: {value!r}")

    return parsed.date()


def closest_past_payment_date(maturity_value, reference_date: date) -> date:
    """
    Find the closest semiannual payment date on or before reference_date.
    """
    maturity_date = parse_any_date(maturity_value)
    maturity_month = maturity_date.month
    maturity_day = maturity_date.day
    previous_month = ((maturity_month - 7) % 12) + 1  # six months earlier

    candidates = []
    for year in (reference_date.year - 1, reference_date.year, reference_date.year + 1):
        candidates.append(_safe_date(year, maturity_month, maturity_day))
        candidates.append(_safe_date(year, previous_month, maturity_day))

    past_candidates = [d for d in candidates if d <= reference_date]
    if not past_candidates:
        raise ValueError(f"No valid past payment date found for maturity {maturity_value!r}")

    return max(past_candidates)


def days_since_last_payment(maturity_value, reference_date: date) -> int:
    last_payment = closest_past_payment_date(maturity_value, reference_date)
    return (reference_date - last_payment).days


def count_payments_until_maturity(maturity_value, reference_date: date) -> int:
    """
    Count semiannual payments from the reference date until maturity,
    INCLUDING the maturity payment itself.
    """
    maturity_date = parse_any_date(maturity_value)
    maturity_month = maturity_date.month
    maturity_day = maturity_date.day
    previous_month = ((maturity_month - 7) % 12) + 1

    payments = []
    for year in range(reference_date.year - 1, maturity_date.year + 2):
        payments.append(_safe_date(year, maturity_month, maturity_day))
        payments.append(_safe_date(year, previous_month, maturity_day))

    valid_payments = sorted(d for d in set(payments) if reference_date < d <= maturity_date)
    return len(valid_payments)


# ============================================================
# BOND PV
# ============================================================

def bond_present_value(
    ask_yield,
    payments_until_maturity,
    coupon_payment,
    face_value=1000,
):
    """
    Bond present value formula.

    Inputs:
        ask_yield               -> annual yield (percent or decimal)
        payments_until_maturity -> number of semiannual periods
        coupon_payment          -> coupon payment per period
        face_value              -> maturity value
    """
    y = float(ask_yield)

    # Convert percentage to decimal if needed
    if y > 1:
        y = y / 100.0

    r = y / 2
    n = int(payments_until_maturity)
    c = float(coupon_payment)
    fv = float(face_value)

    if n <= 0:
        return fv

    if abs(r) < 1e-12:
        return (c * n) + fv

    pv_coupons = c * (1 - (1 + r) ** (-n)) / r
    pv_face = fv / ((1 + r) ** n)
    return pv_coupons + pv_face

def bond_future_value(
    ask_yield,
    payments_until_maturity,
    coupon_payment,
    face_value=1000,
):
    """
    Bond present value formula.

    Inputs:
        ask_yield               -> annual yield (percent or decimal)
        payments_until_maturity -> number of semiannual periods
        coupon_payment          -> coupon payment per period
        face_value              -> maturity value
    """
    y = float(ask_yield)

    # Convert percentage to decimal if needed
    y+=1
    if y > 1:
        y = y / 100.0

    r = y / 2
    n = int(payments_until_maturity)
    c = float(coupon_payment)
    fv = float(face_value)

    if n <= 0:
        return fv
    if n<=2:
        n=0
    if n>=3:
        n-=2
    if abs(r) < 1e-12:
        return (c * n) + fv

    pv_coupons = c * (1 - (1 + r) ** (-n)) / r
    pv_face = fv / ((1 + r) ** n)
    return pv_coupons + pv_face


# ============================================================
# REFERENCE DATE EXTRACTION
# ============================================================

def parse_date_string(value: str) -> Optional[date]:
    if not value:
        return None

    parsed = pd.to_datetime(value, errors="coerce", utc=True)
    if pd.isna(parsed):
        parsed = pd.to_datetime(value, errors="coerce")

    if pd.isna(parsed):
        return None

    return parsed.date()


def first_regex_match(patterns: Sequence[str], text: str) -> Optional[str]:
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        if m:
            return m.group(1).strip()
    return None


def parse_reference_date_from_html(html: str) -> Optional[date]:
    patterns = [
        r'<meta[^>]+property=["\']article:published_time["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+name=["\']article:published_time["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+property=["\']article:modified_time["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+name=["\']article:modified_time["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+name=["\']parsely-pub-date["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+property=["\']og:updated_time["\'][^>]+content=["\']([^"\']+)["\']',
        r'"datePublished"\s*:\s*"([^"]+)"',
        r'"dateModified"\s*:\s*"([^"]+)"',
        r'<time[^>]+datetime=["\']([^"\']+)["\']',
        r"\bPublished\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})\b",
        r"\bUpdated\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})\b",
        r"\bAs of\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})\b",
    ]

    raw = first_regex_match(patterns, html)
    return parse_date_string(raw) if raw else None


def get_reference_date(url: str, explicit_reference_date: Optional[str] = None) -> date:
    if explicit_reference_date:
        parsed = parse_date_string(explicit_reference_date)
        if parsed is None:
            raise ValueError(f"Could not parse --reference-date: {explicit_reference_date!r}")
        return parsed

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
    }

    try:
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        parsed = parse_reference_date_from_html(resp.text)
        if parsed is not None:
            return parsed
    except Exception:
        pass

    # Playwright fallback for date metadata
    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(user_agent=headers["User-Agent"], viewport={"width": 1600, "height": 1400})
            page.goto(url, wait_until="domcontentloaded", timeout=120_000)
            page.wait_for_timeout(5000)

            for selector in [
                "meta[property='article:published_time']",
                "meta[name='article:published_time']",
                "meta[property='article:modified_time']",
                "meta[name='article:modified_time']",
                "meta[name='parsely-pub-date']",
                "meta[property='og:updated_time']",
            ]:
                loc = page.locator(selector)
                if loc.count() > 0:
                    content = loc.first.get_attribute("content")
                    parsed = parse_date_string(content) if content else None
                    if parsed is not None:
                        browser.close()
                        return parsed

            html = page.content()
            browser.close()
            parsed = parse_reference_date_from_html(html)
            if parsed is not None:
                return parsed
    except Exception:
        pass

    today = date.today()
    print(
        f"Warning: could not determine article date. Using today's date instead: {today.isoformat()}",
        file=sys.stderr,
    )
    return today


# ============================================================
# TABLE / JSON EXTRACTION HELPERS
# ============================================================

def normalize_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name).strip().lower())


def normalize_target_headers(headers: Sequence[str]) -> List[str]:
    out = []
    for h in headers:
        n = normalize_col(h)
        if n in {"askedyield", "askyield", "askyld"}:
            out.append("Asked Yield")
        elif n == "maturity":
            out.append("Maturity")
        elif n == "coupon":
            out.append("Coupon")
        else:
            out.append(str(h).strip())
    return out


def filter_target_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {}
    for col in df.columns:
        n = normalize_col(col)
        if n == normalize_col("Maturity"):
            rename_map[col] = "Maturity"
        elif n == normalize_col("Coupon"):
            rename_map[col] = "Coupon"
        elif n in {
            normalize_col("Asked Yield"),
            normalize_col("Ask Yield"),
            normalize_col("AskYld"),
        }:
            rename_map[col] = "Asked Yield"

    df = df.rename(columns=rename_map)
    keep = [c for c in TARGET_COLUMNS if c in df.columns]
    if len(keep) != 3:
        return pd.DataFrame(columns=TARGET_COLUMNS)

    out = df[keep].copy()
    out.columns = TARGET_COLUMNS
    return out


def table_from_rows(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> pd.DataFrame:
    headers = normalize_target_headers(headers)
    data = []
    for row in rows:
        row = [str(c).strip() for c in row]
        if len(row) < len(headers):
            row = list(row) + [""] * (len(headers) - len(row))
        elif len(row) > len(headers):
            row = list(row[: len(headers)])
        data.append(row)
    return pd.DataFrame(data, columns=headers)


def extract_records_from_obj(obj: Any) -> List[dict]:
    found: List[dict] = []

    def walk(x: Any):
        if isinstance(x, dict):
            norm_keys = {normalize_col(k): k for k in x.keys()}
            has_maturity = normalize_col("Maturity") in norm_keys
            has_coupon = normalize_col("Coupon") in norm_keys
            has_asked = (
                normalize_col("Asked Yield") in norm_keys
                or normalize_col("Ask Yield") in norm_keys
            )

            if has_maturity and has_coupon and has_asked:
                found.append(x)

            for v in x.values():
                walk(v)

        elif isinstance(x, list):
            for item in x:
                walk(item)

    walk(obj)
    return found


def records_to_df(records: List[dict]) -> pd.DataFrame:
    rows = []
    for rec in records:
        row = {}
        for key, val in rec.items():
            nk = normalize_col(key)
            if nk == normalize_col("Maturity"):
                row["Maturity"] = val
            elif nk == normalize_col("Coupon"):
                row["Coupon"] = val
            elif nk in {
                normalize_col("Asked Yield"),
                normalize_col("Ask Yield"),
                normalize_col("AskYld"),
            }:
                row["Asked Yield"] = val
        if len(row) == 3:
            rows.append(row)

    if not rows:
        return pd.DataFrame(columns=TARGET_COLUMNS)

    df = pd.DataFrame(rows)
    return df[TARGET_COLUMNS]


def try_parse_json_text(text: str) -> Optional[pd.DataFrame]:
    text = text.strip()
    if not text:
        return None

    # Direct JSON
    try:
        obj = json.loads(text)
        records = extract_records_from_obj(obj)
        if records:
            return records_to_df(records)
    except Exception:
        pass

    # Regex-based JSON fragment search
    patterns = [
        r'(\{[^{}]{0,4000}?"Maturity"[^{}]{0,4000}?"Coupon"[^{}]{0,4000}?"Asked Yield"[^{}]{0,4000}?\})',
        r'(\{[^{}]{0,4000}?"Maturity"[^{}]{0,4000}?"Coupon"[^{}]{0,4000}?"Ask Yield"[^{}]{0,4000}?\})',
    ]
    for pat in patterns:
        for m in re.finditer(pat, text, flags=re.IGNORECASE | re.DOTALL):
            frag = m.group(1)
            try:
                obj = json.loads(frag)
                records = extract_records_from_obj(obj)
                if records:
                    return records_to_df(records)
            except Exception:
                continue

    return None


def parse_html_tables(html: str) -> Optional[pd.DataFrame]:
    try:
        tables = pd.read_html(html)
    except Exception:
        return None

    for table in tables:
        out = filter_target_columns(table)
        if not out.empty:
            return out
    return None


def fetch_with_requests(url: str) -> Optional[pd.DataFrame]:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
    }

    resp = requests.get(url, headers=headers, timeout=30)
    resp.raise_for_status()
    html = resp.text

    df = parse_html_tables(html)
    if df is not None:
        return df

    # Try script tags
    script_matches = re.findall(
        r"<script[^>]*>(.*?)</script>",
        html,
        flags=re.DOTALL | re.IGNORECASE,
    )
    for script_text in script_matches:
        df = try_parse_json_text(script_text)
        if df is not None and not df.empty:
            return df

    # Try whole HTML as a JSON container if the page ships embedded state
    df = try_parse_json_text(html)
    if df is not None and not df.empty:
        return df

    return None


def extract_with_playwright(url: str) -> Optional[pd.DataFrame]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return None

    user_agent = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )

    collected_texts: List[str] = []

    def on_response(resp):
        try:
            ct = (resp.headers.get("content-type") or "").lower()
            if "json" in ct or "text" in ct or "javascript" in ct:
                txt = resp.text()
                if txt:
                    collected_texts.append(txt)
        except Exception:
            pass

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(user_agent=user_agent, viewport={"width": 1600, "height": 1400})
        page.on("response", on_response)

        page.goto(url, wait_until="domcontentloaded", timeout=120_000)
        page.wait_for_timeout(7000)

        # 1) Standard tables/grids
        for selector in ["table", "[role='table']", "[role='grid']"]:
            count = page.locator(selector).count()
            for i in range(count):
                root = page.locator(selector).nth(i)
                headers = [
                    t.strip()
                    for t in root.locator("th, [role='columnheader']").all_inner_texts()
                    if t.strip()
                ]
                if not headers:
                    first_row = root.locator("tr, [role='row']").first
                    headers = [
                        t.strip()
                        for t in first_row.locator(
                            "th, td, [role='columnheader'], [role='gridcell']"
                        ).all_inner_texts()
                        if t.strip()
                    ]
                if not headers:
                    continue

                rows = []
                for row in root.locator("tr, [role='row']").all()[1:]:
                    cells = [
                        t.strip()
                        for t in row.locator("th, td, [role='gridcell']").all_inner_texts()
                    ]
                    cells = [c for c in cells if c != ""]
                    if cells:
                        rows.append(cells)

                if rows:
                    df = table_from_rows(headers, rows)
                    out = filter_target_columns(df)
                    if not out.empty:
                        browser.close()
                        return out

        # 2) Visible page text
        body_text = page.locator("body").inner_text(timeout=10_000)
        html = page.content()
        browser.close()

    # 3) Parse rendered HTML
    df = parse_html_tables(html)
    if df is not None:
        return df

    # 4) Parse embedded JSON/state in rendered HTML
    df = try_parse_json_text(html)
    if df is not None and not df.empty:
        return df

    # 5) Parse JSON captured from network responses
    for txt in collected_texts:
        df = try_parse_json_text(txt)
        if df is not None and not df.empty:
            return df

    # 6) Very last resort: reconstruct from visible lines
    lines = [re.sub(r"\s+", " ", ln).strip() for ln in body_text.splitlines()]
    lines = [ln for ln in lines if ln]

    header_idx = None
    for i, line in enumerate(lines):
        low = line.lower()
        if "maturity" in low and "coupon" in low and "asked yield" in low:
            header_idx = i
            break

    if header_idx is None:
        return None

    data_lines = lines[header_idx + 1:]
    rows = []
    for line in data_lines:
        parts = [p.strip() for p in re.split(r"\s{2,}", line) if p.strip()]
        if len(parts) >= 3:
            rows.append(parts[:3])

    if rows:
        return pd.DataFrame(rows, columns=TARGET_COLUMNS)

    return None


# ============================================================
# BARCHART SCRAPER
# ============================================================
#
# Strategy: Barchart's page is a React SPA that actively blocks headless
# browsers and does not render a parseable table in static HTML.
#
# Instead we:
#  1. Open a real requests.Session and hit the homepage to collect cookies
#     and the XSRF token that Barchart requires on every API call.
#  2. Call their internal /proxima/core/getQuote endpoint (same one the
#     browser's XHR calls) with the well-known financial futures symbols.
#  3. If that fails, try their legacy marketdata.websol endpoint with a
#     free guest key that Barchart embeds in every page.
#  4. If both fail, fall back to Playwright with network interception so
#     we can capture the real XHR response.
#
# Known financial futures symbols on Barchart (front-month *0 notation):
BARCHART_FIN_SYMBOLS = [
    "ZT*0",   # 2-Year T-Note
    "ZF*0",   # 5-Year T-Note
    "ZN*0",   # 10-Year T-Note
    "ZB*0",   # 30-Year T-Bond
    "UB*0",   # Ultra T-Bond
    "GE*0",   # Eurodollar (CME)
    "SR3*0",  # SOFR 3-Month
    "FF*0",   # 30-Day Fed Funds
    "ZQ*0",   # 30-Day Fed Funds (alt)
    "MES*0",  # Micro E-mini S&P (included on some views)
]

_BC_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.barchart.com/futures/financials?viewName=main",
    "Origin": "https://www.barchart.com",
}


def _normalise_barchart_headers(headers: Sequence[str]) -> List[str]:
    """Map raw header text to canonical BARCHART_COLUMNS names."""
    mapping = {
        "symbol":           "Symbol",
        "contractname":     "Contract Name",
        "contract":         "Contract Name",
        "name":             "Contract Name",
        "contractmonth":    "Contract Name",
        "last":             "Latest",
        "lastprice":        "Latest",
        "latest":           "Latest",
        "price":            "Latest",
        "change":           "Change",
        "chg":              "Change",
        "changepercent":    "Change",
        "volume":           "Volume",
        "vol":              "Volume",
        "time":             "Time",
        "updated":          "Time",
        "lastupdated":      "Time",
        "tradetime":        "Time",
        "timestamp":        "Time",
    }
    out = []
    for h in headers:
        key = re.sub(r"[^a-z0-9]+", "", str(h).strip().lower())
        out.append(mapping.get(key, str(h).strip()))
    return out


def _barchart_df_from_rows(
    headers: Sequence[str], rows: Sequence[Sequence[str]]
) -> Optional[pd.DataFrame]:
    norm_headers = _normalise_barchart_headers(headers)
    if not any(h in BARCHART_COLUMNS for h in norm_headers):
        return None

    data = []
    for row in rows:
        row = [str(c).strip() for c in row]
        if len(row) < len(norm_headers):
            row = list(row) + [""] * (len(norm_headers) - len(row))
        elif len(row) > len(norm_headers):
            row = list(row[: len(norm_headers)])
        data.append(row)

    if not data:
        return None

    df = pd.DataFrame(data, columns=norm_headers)
    keep = [c for c in BARCHART_COLUMNS if c in df.columns]
    if not keep:
        return None

    out = df[keep].copy()
    if "Symbol" in out.columns:
        out = out[out["Symbol"].str.strip() != ""]
        out = out[~out["Symbol"].str.lower().isin({"symbol", "contract"})]
    return out if not out.empty else None


def _barchart_try_api_json(collected_jsons: List[str]) -> Optional[pd.DataFrame]:
    """
    Barchart's page calls internal REST endpoints that return JSON arrays
    shaped like:
      {"data": [{"symbol":"ZT*0","contractName":"2-Year T-Note","last":"...","change":"...","volume":"...","tradeTime":"..."}, ...]}
    or lists of such dicts at the top level.
    We scan every captured JSON payload looking for these shapes.
    """
    key_aliases: dict[str, str] = {
        # symbol
        "symbol": "Symbol",
        "ticker": "Symbol",
        # contract name
        "contractname": "Contract Name",
        "contract": "Contract Name",
        "name": "Contract Name",
        "description": "Contract Name",
        # latest / last price
        "last": "Latest",
        "lastprice": "Latest",
        "latest": "Latest",
        "close": "Latest",
        "price": "Latest",
        # change
        "change": "Change",
        "chg": "Change",
        "pricechange": "Change",
        "netchange": "Change",
        # volume
        "volume": "Volume",
        "vol": "Volume",
        # time
        "tradetime": "Time",
        "timestamp": "Time",
        "time": "Time",
        "lasttradetime": "Time",
        "updated": "Time",
    }

    def _records_from_obj(obj: Any) -> List[dict]:
        """Recursively find dicts that contain at least Symbol + one other target."""
        found: List[dict] = []

        def walk(x: Any) -> None:
            if isinstance(x, dict):
                norm = {re.sub(r"[^a-z0-9]", "", k.lower()): v for k, v in x.items()}
                has_sym = "symbol" in norm or "ticker" in norm
                has_price = any(k in norm for k in ("last", "lastprice", "latest", "close", "price"))
                if has_sym and has_price:
                    row: dict = {}
                    for raw_k, val in x.items():
                        canon = key_aliases.get(re.sub(r"[^a-z0-9]", "", raw_k.lower()))
                        if canon:
                            row.setdefault(canon, str(val).strip())
                    if row.get("Symbol"):
                        found.append(row)
                for v in x.values():
                    walk(v)
            elif isinstance(x, list):
                for item in x:
                    walk(item)

        walk(obj)
        return found

    for text in collected_jsons:
        text = text.strip()
        if not text or text[0] not in ("{", "["):
            continue
        try:
            obj = json.loads(text)
        except Exception:
            # Try to pull JSON fragments
            for m in re.finditer(r'(\{[^{}]{10,8000}\})', text):
                try:
                    obj = json.loads(m.group(1))
                    records = _records_from_obj(obj)
                    if records:
                        df = pd.DataFrame(records)
                        for col in BARCHART_COLUMNS:
                            if col not in df.columns:
                                df[col] = ""
                        return df[BARCHART_COLUMNS].copy()
                except Exception:
                    continue
            continue

        records = _records_from_obj(obj)
        if records:
            df = pd.DataFrame(records)
            for col in BARCHART_COLUMNS:
                if col not in df.columns:
                    df[col] = ""
            return df[BARCHART_COLUMNS].copy()

    return None


def _barchart_parse_dom(page: Any, log: Any = None) -> Optional[pd.DataFrame]:
    """
    Try every plausible DOM strategy to pull the Barchart data table.
    `page` is a Playwright Page object.
    `log` is an optional callable(str) for debug messages.
    """
    def _dbg(msg: str) -> None:
        if log:
            log(msg)

    # ── Strategy A: Barchart's own bc-table / data-ng-* component ────────────
    # Barchart renders a custom <table> inside a .bc-table wrapper.
    # The symbol column uses <span class="symbol"> or an <a> tag.
    # We try to pull rows by targeting the inner table first.
    for tbl_sel in [
        "table.bc-futures-financials-table",
        "table.bc-datatable",
        ".bc-table-scrollable-inner table",
        ".bc-datatable-wrapper table",
        "table",
    ]:
        try:
            count = page.locator(tbl_sel).count()
            _dbg(f"Selector '{tbl_sel}': {count} element(s)")
            for i in range(count):
                root = page.locator(tbl_sel).nth(i)

                # Headers: try <thead th> first, then first <tr> cells
                raw_hdrs: List[str] = []
                thead = root.locator("thead th, thead td")
                if thead.count() > 0:
                    raw_hdrs = [t.strip() for t in thead.all_inner_texts() if t.strip()]
                if not raw_hdrs:
                    first_tr = root.locator("tr").first
                    raw_hdrs = [
                        t.strip()
                        for t in first_tr.locator("th, td").all_inner_texts()
                        if t.strip()
                    ]
                if not raw_hdrs:
                    continue

                _dbg(f"  table[{i}] headers: {raw_hdrs}")

                # Rows: every <tr> in <tbody>, or all <tr> after the first
                rows_data: List[List[str]] = []
                tbody_rows = root.locator("tbody tr")
                if tbody_rows.count() > 0:
                    row_els = tbody_rows.all()
                else:
                    all_trs = root.locator("tr").all()
                    row_els = all_trs[1:]  # skip header row

                for row_el in row_els:
                    cells = [t.strip() for t in row_el.locator("td, th").all_inner_texts()]
                    cells = [c for c in cells if c]
                    if cells:
                        rows_data.append(cells)

                _dbg(f"  table[{i}] data rows: {len(rows_data)}")
                if rows_data:
                    df = _barchart_df_from_rows(raw_hdrs, rows_data)
                    if df is not None and not df.empty:
                        _dbg(f"  ✓ Extracted {len(df)} rows via '{tbl_sel}'")
                        return df
        except Exception as exc:
            _dbg(f"  Error with '{tbl_sel}': {exc}")
            continue

    # ── Strategy B: generic [role='table'] / [role='grid'] ───────────────────
    for grid_sel in ["[role='table']", "[role='grid']"]:
        try:
            count = page.locator(grid_sel).count()
            _dbg(f"Grid selector '{grid_sel}': {count}")
            for i in range(count):
                root = page.locator(grid_sel).nth(i)
                raw_hdrs = [
                    t.strip()
                    for t in root.locator("[role='columnheader']").all_inner_texts()
                    if t.strip()
                ]
                if not raw_hdrs:
                    continue
                _dbg(f"  grid[{i}] headers: {raw_hdrs}")
                rows_data = []
                for row_el in root.locator("[role='row']").all()[1:]:
                    cells = [
                        t.strip()
                        for t in row_el.locator("[role='cell'], [role='gridcell']").all_inner_texts()
                    ]
                    cells = [c for c in cells if c]
                    if cells:
                        rows_data.append(cells)
                _dbg(f"  grid[{i}] data rows: {len(rows_data)}")
                if rows_data:
                    df = _barchart_df_from_rows(raw_hdrs, rows_data)
                    if df is not None and not df.empty:
                        _dbg(f"  ✓ Extracted {len(df)} rows via grid")
                        return df
        except Exception as exc:
            _dbg(f"  Error with '{grid_sel}': {exc}")

    # ── Strategy C: page-level inner_text line reconstruction ────────────────
    # Barchart rows all contain a ticker-like symbol (e.g. "ZT*0") and a price.
    # We try to reconstruct rows from visible body text.
    try:
        body_text = page.locator("body").inner_text(timeout=15_000)
        lines = [re.sub(r"\s{2,}", "\t", ln).strip() for ln in body_text.splitlines()]
        lines = [ln for ln in lines if ln]

        # Find header line containing "Symbol" and at least one other target word
        header_idx: Optional[int] = None
        for idx, line in enumerate(lines):
            low = line.lower()
            if ("symbol" in low or "contract" in low) and (
                "last" in low or "change" in low or "volume" in low
            ):
                header_idx = idx
                break

        if header_idx is not None:
            raw_hdrs = [p.strip() for p in lines[header_idx].split("\t") if p.strip()]
            _dbg(f"Text-recon header: {raw_hdrs}")
            rows_data = []
            for line in lines[header_idx + 1:]:
                parts = [p.strip() for p in line.split("\t") if p.strip()]
                if len(parts) >= 2:
                    rows_data.append(parts)
            _dbg(f"Text-recon rows: {len(rows_data)}")
            if rows_data:
                df = _barchart_df_from_rows(raw_hdrs, rows_data)
                if df is not None and not df.empty:
                    _dbg(f"✓ Extracted {len(df)} rows via text reconstruction")
                    return df
    except Exception as exc:
        _dbg(f"Text recon error: {exc}")

    return None


def _barchart_with_playwright(url: str, log: Any = None) -> Optional[pd.DataFrame]:
    """
    Full Playwright-based Barchart scraper.
    1. Navigates to the page and waits for JS to render.
    2. Intercepts all XHR/fetch responses to catch API JSON payloads.
    3. Dismisses cookie banners.
    4. Tries DOM extraction strategies.
    5. Falls back to captured JSON from network.
    6. Falls back to pd.read_html on the rendered HTML.
    """
    def _dbg(msg: str) -> None:
        if log:
            log(msg)

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        _dbg("Playwright not installed.")
        return None

    user_agent = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )

    captured_jsons: List[str] = []

    def _on_response(resp: Any) -> None:
        try:
            ct = (resp.headers.get("content-type") or "").lower()
            req_url = resp.url.lower()
            # Capture JSON and JS responses; focus on API endpoints
            if "json" in ct or ("javascript" in ct and "getquotes" in req_url):
                body = resp.text()
                if body and len(body) > 50:
                    captured_jsons.append(body)
        except Exception:
            pass

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--disable-blink-features=AutomationControlled"])
        ctx = browser.new_context(
            user_agent=user_agent,
            viewport={"width": 1600, "height": 1200},
            locale="en-US",
        )
        page = ctx.new_page()
        page.set_extra_http_headers({
            "Referer": "https://www.barchart.com/",
            "Accept-Language": "en-US,en;q=0.9",
        })
        page.on("response", _on_response)

        _dbg(f"Navigating to {url}")
        try:
            page.goto(url, wait_until="networkidle", timeout=90_000)
        except Exception:
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=90_000)
            except Exception as e:
                _dbg(f"Navigation error: {e}")

        # Give JS time to render
        page.wait_for_timeout(6000)

        # Dismiss cookie / consent overlays
        for btn_text in ["Accept All", "Accept", "I Agree", "Got it", "OK", "Close"]:
            try:
                btn = page.get_by_role("button", name=re.compile(btn_text, re.I))
                if btn.first.is_visible(timeout=800):
                    btn.first.click()
                    page.wait_for_timeout(1000)
                    _dbg(f"Dismissed overlay: '{btn_text}'")
            except Exception:
                pass

        # Extra wait after dismissals for data to render
        page.wait_for_timeout(4000)

        # Log what's on the page for diagnostics
        _dbg(f"Captured {len(captured_jsons)} JSON responses from network")

        # ── Try DOM first ──────────────────────────────────────────────────────
        df = _barchart_parse_dom(page, log=log)
        if df is not None and not df.empty:
            browser.close()
            return df

        # ── Try network JSON ───────────────────────────────────────────────────
        _dbg("DOM strategies exhausted; trying captured network JSON...")
        df = _barchart_try_api_json(captured_jsons)
        if df is not None and not df.empty:
            _dbg(f"✓ Extracted {len(df)} rows from network JSON")
            browser.close()
            return df

        # ── Fallback: pd.read_html on rendered HTML ────────────────────────────
        _dbg("Trying pd.read_html on rendered HTML...")
        html = page.content()
        browser.close()

    try:
        tables = pd.read_html(html)
        _dbg(f"pd.read_html found {len(tables)} table(s)")
        for i, table in enumerate(tables):
            normed_cols = _normalise_barchart_headers(list(table.columns.astype(str)))
            table.columns = normed_cols
            keep = [c for c in BARCHART_COLUMNS if c in table.columns]
            _dbg(f"  table[{i}] matched columns: {keep}")
            if len(keep) >= 3:
                out = table[keep].copy()
                if "Symbol" in out.columns:
                    out = out[out["Symbol"].astype(str).str.strip() != ""]
                if not out.empty:
                    _dbg(f"  ✓ Extracted {len(out)} rows from rendered HTML table[{i}]")
                    return out
    except Exception as exc:
        _dbg(f"pd.read_html error: {exc}")

    return None


def _barchart_session() -> requests.Session:
    """Return a requests.Session pre-loaded with Barchart cookies + XSRF token."""
    s = requests.Session()
    s.headers.update(_BC_HEADERS)
    try:
        r = s.get("https://www.barchart.com/futures/financials?viewName=main", timeout=30)
        xsrf = s.cookies.get("XSRF-TOKEN", "")
        if xsrf:
            s.headers["X-XSRF-TOKEN"] = requests.utils.unquote(xsrf)
    except Exception:
        pass
    return s


def _records_to_df(records: list, log: Any = None) -> Optional[pd.DataFrame]:
    """Convert a list of dicts (from any Barchart API) into a BARCHART_COLUMNS DataFrame.

    Barchart returns camelCase keys like: symbol, contractName, lastPrice,
    priceChange, netChange, volume, tradeTime.
    We normalise by stripping all non-alpha chars and lower-casing before lookup.
    """
    key_map = {
        # Symbol
        "symbol":           "Symbol",
        "ticker":           "Symbol",
        # Contract Name
        "contractname":     "Contract Name",
        "contract":         "Contract Name",
        "name":             "Contract Name",
        "description":      "Contract Name",
        # Latest — Barchart uses "lastPrice" -> stripped "lastprice"
        "lastprice":        "Latest",
        "last":             "Latest",
        "lasttradeprice":   "Latest",
        "close":            "Latest",
        "closeprice":       "Latest",
        "price":            "Latest",
        "latest":           "Latest",
        # Change — Barchart uses "priceChange" or "netChange"
        "pricechange":      "Change",
        "netchange":        "Change",
        "change":           "Change",
        "chg":              "Change",
        "changepercent":    "Change",
        # Volume
        "volume":           "Volume",
        "vol":              "Volume",
        # Time — Barchart uses "tradeTime"
        "tradetime":        "Time",
        "lasttradetime":    "Time",
        "timestamp":        "Time",
        "time":             "Time",
        "updated":          "Time",
        "lastupdated":      "Time",
    }

    def _norm(k: str) -> str:
        return re.sub(r"[^a-z]", "", k.lower())

    rows = []
    for rec in records:
        if not isinstance(rec, dict):
            continue
        row: dict = {}
        for k, v in rec.items():
            canon = key_map.get(_norm(k))
            if canon:
                row.setdefault(canon, str(v).strip() if v is not None else "")
        # If Change is still missing, do a second pass looking for any key
        # containing "change" that has a numeric-looking value
        if row.get("Symbol") and not row.get("Change"):
            for k, v in rec.items():
                if "change" in k.lower() and v not in (None, "", "0", 0):
                    row.setdefault("Change", str(v).strip())
                    break
        if row.get("Symbol"):
            rows.append(row)

    if log and rows:
        log(f"  Sample mapped row: {dict(list(rows[0].items()))}")

    if not rows:
        return None
    df = pd.DataFrame(rows)
    for col in BARCHART_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[BARCHART_COLUMNS].copy()


def _barchart_via_proxima(session: requests.Session, log: Any = None) -> Optional[pd.DataFrame]:
    """
    Call Barchart's internal /proxima/core/getQuote API.
    This is the same XHR the browser page fires.
    """
    def _dbg(m: str) -> None:
        if log: log(m)

    symbols = ",".join(BARCHART_FIN_SYMBOLS)
    fields = "symbol,contractName,lastPrice,priceChange,netChange,change,volume,tradeTime"
    url = (
        f"https://www.barchart.com/proxima/core/getQuote"
        f"?symbols={requests.utils.quote(symbols)}&fields={fields}"
        f"&raw=1"
    )
    _dbg(f"Trying proxima API: {url}")
    try:
        r = session.get(url, timeout=30)
        _dbg(f"  Status: {r.status_code}")
        if r.status_code == 200:
            data = r.json()
            # Response shape: {"data": [...]} or [...]
            records = data.get("data", data) if isinstance(data, dict) else data
            if isinstance(records, list) and records:
                _dbg(f"  Raw keys in first record: {list(records[0].keys()) if records else []}")
                df = _records_to_df(records, log=log)
                if df is not None and not df.empty:
                    _dbg(f"  ✓ proxima returned {len(df)} rows")
                    return df
    except Exception as e:
        _dbg(f"  proxima error: {e}")
    return None


def _barchart_via_marketdata(session: requests.Session, log: Any = None) -> Optional[pd.DataFrame]:
    """
    Call the legacy marketdata.websol.barchart.com getQuote endpoint.
    Barchart embeds a free guest API key in their pages; we use the
    well-known public key that works for delayed quotes.
    """
    def _dbg(m: str) -> None:
        if log: log(m)

    # Barchart's publicly-embedded free key (delayed data, no sign-up needed)
    api_key = "bc4e17d7e8f3447b947cf9f2a3c6c8f2"
    symbols = ",".join(BARCHART_FIN_SYMBOLS)
    url = (
        f"https://marketdata.websol.barchart.com/getQuote.json"
        f"?apikey={api_key}&symbols={requests.utils.quote(symbols)}"
        f"&fields=contractName,lastPrice,priceChange,netChange,change,volume,tradeTime"
    )
    _dbg(f"Trying marketdata API: {url}")
    try:
        r = session.get(url, timeout=30)
        _dbg(f"  Status: {r.status_code}")
        if r.status_code == 200:
            data = r.json()
            records = data.get("results", [])
            if records:
                _dbg(f"  Raw keys in first record: {list(records[0].keys()) if records else []}")
                df = _records_to_df(records, log=log)
                if df is not None and not df.empty:
                    _dbg(f"  ✓ marketdata returned {len(df)} rows")
                    return df
    except Exception as e:
        _dbg(f"  marketdata error: {e}")
    return None


def _barchart_via_playwright_intercept(url: str, log: Any = None) -> Optional[pd.DataFrame]:
    """
    Full Playwright fallback: navigate with a real browser, intercept every
    JSON XHR, and scan payloads for quote records.
    """
    def _dbg(m: str) -> None:
        if log: log(m)

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        _dbg("Playwright not installed — skipping browser fallback.")
        return None

    captured: List[str] = []

    def on_response(resp: Any) -> None:
        try:
            ct = resp.headers.get("content-type", "")
            if "json" in ct:
                body = resp.text()
                if body and len(body) > 80:
                    captured.append(body)
        except Exception:
            pass

    _dbg("Launching Playwright browser…")
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled"],
        )
        ctx = browser.new_context(
            user_agent=_BC_HEADERS["User-Agent"],
            viewport={"width": 1600, "height": 1200},
            locale="en-US",
        )
        page = ctx.new_page()
        page.set_extra_http_headers({
            "Referer": "https://www.barchart.com/",
            "Accept-Language": "en-US,en;q=0.9",
        })
        page.on("response", on_response)

        try:
            page.goto(url, wait_until="networkidle", timeout=90_000)
        except Exception:
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=60_000)
                page.wait_for_timeout(10_000)
            except Exception as e:
                _dbg(f"Navigation error: {e}")

        page.wait_for_timeout(5_000)
        _dbg(f"Captured {len(captured)} JSON payloads from network")

        # Grab XSRF token from browser cookies, then retry the proxima API
        cookies = ctx.cookies()
        xsrf = next((c["value"] for c in cookies if c["name"] == "XSRF-TOKEN"), None)
        if xsrf:
            _dbg("Got XSRF token from browser — retrying proxima API with real session…")
            session = requests.Session()
            session.headers.update(_BC_HEADERS)
            session.headers["X-XSRF-TOKEN"] = requests.utils.unquote(xsrf)
            for c in cookies:
                session.cookies.set(c["name"], c["value"], domain=c.get("domain", ".barchart.com"))
            browser.close()
            df = _barchart_via_proxima(session, log=log)
            if df is not None and not df.empty:
                return df
            df = _barchart_via_marketdata(session, log=log)
            if df is not None and not df.empty:
                return df
        else:
            browser.close()

        # Scan captured JSON payloads
        key_map = {
            "symbol": "Symbol", "contractname": "Contract Name",
            "name": "Contract Name", "description": "Contract Name",
            "last": "Latest", "lastprice": "Latest", "close": "Latest",
            "change": "Change", "netchange": "Change",
            "volume": "Volume",
            "tradetime": "Time", "time": "Time", "lasttradetime": "Time",
        }

        def _walk(obj: Any, out: list) -> None:
            if isinstance(obj, dict):
                norm = {re.sub(r"[^a-z0-9]", "", k.lower()): v for k, v in obj.items()}
                if ("symbol" in norm or "contractname" in norm) and (
                    "lastprice" in norm or "last" in norm or "close" in norm
                ):
                    row: dict = {}
                    for k, v in obj.items():
                        canon = key_map.get(re.sub(r"[^a-z]", "", k.lower()))
                        if canon:
                            row.setdefault(canon, str(v).strip() if v is not None else "")
                    # Fallback: grab ANY key containing "change" with a non-zero value
                    if row.get("Symbol") and not row.get("Change"):
                        for k, v in obj.items():
                            if "change" in k.lower() and v not in (None, "", "0", 0):
                                row["Change"] = str(v).strip()
                                break
                    # Log all raw keys on first record so we can see exact field names
                    if row.get("Symbol") and not out:
                        _dbg(f"  Raw payload keys: {list(obj.keys())}")
                        _dbg(f"  Raw payload values: { {k: v for k, v in list(obj.items())[:12]} }")
                    if row.get("Symbol"):
                        out.append(row)
                for v in obj.values():
                    _walk(v, out)
            elif isinstance(obj, list):
                for item in obj:
                    _walk(item, out)

        all_records: list = []
        for payload in captured:
            try:
                obj = json.loads(payload)
                _walk(obj, all_records)
            except Exception:
                pass

        _dbg(f"Found {len(all_records)} quote records in network payloads")
        if all_records:
            df = _records_to_df(all_records, log=log)
            if df is not None and not df.empty:
                _dbg(f"✓ Extracted {len(df)} rows from intercepted network data")
                return df

    return None


def _convert_treasury_price(val: str) -> str:
    """
    Convert a Barchart treasury price string to a decimal.

    Format rules:
      XX-YY      → XX + YY/32
      XX-YYZ     → XX + YY/32 + Z/8/32   (Z is a single digit: 2=¼, 4=½, 6=¾)
      XX-YY+     → XX + YY/32 + 0.5/32   ('+' means half a 32nd)
      Trailing letters (e.g. 's' for "settlement") are stripped before parsing.
      Anything not matching is returned unchanged.

    Examples:
      "97-08"    → 97.25        (8/32 = 0.25)
      "111-22"   → 111.6875     (22/32 = 0.6875)
      "109-162"  → 109.515625   (16/32 + 2/8/32)
      "97-08+"   → 97.265625    (8/32 + 0.5/32)
      "113-23s"  → 113.71875    (23/32, trailing 's' stripped)
      "109-305s" → 109.957031   (30/32 + 5/8/32, trailing 's' stripped)
    """
    if not isinstance(val, str):
        return str(val)
    s = val.strip()

    # Handle leading +/- sign for change column
    sign = 1
    if s.startswith(("+", "-")):
        sign = -1 if s[0] == "-" else 1
        s = s[1:].strip()

    # Strip trailing letter flags (e.g. Barchart's 's' for "settlement")
    s = re.sub(r"[a-zA-Z]+$", "", s)

    # Match: whole-YY or whole-YYZ or whole-YY+
    m = re.match(r"^(\d+)-(\d{2,3})(\+?)$", s)
    if not m:
        return val  # not a treasury price format — leave as-is

    whole   = int(m.group(1))
    frac_s  = m.group(2)
    plus    = m.group(3) == "+"

    if len(frac_s) == 2:
        # XX-YY
        thirty_seconds = int(frac_s)
        extra = 0.0
    else:
        # XX-YYZ  where Z is the third digit (eighths of a 32nd)
        thirty_seconds = int(frac_s[:2])
        extra = int(frac_s[2]) / 8.0

    decimal = whole + (thirty_seconds + extra + (0.5 if plus else 0.0)) / 32.0
    result  = f"{sign * decimal:.6f}".rstrip("0").rstrip(".")
    return result


def _postprocess_barchart(df: pd.DataFrame, log: Any = None) -> pd.DataFrame:
    """
    Clean up the raw Barchart dataframe:
      1. Drop exact duplicate rows (same Symbol + Latest).
      2. Drop rows that are just repeated header/section rows.
      3. Convert Latest and Change from treasury fractional notation to decimals.
    """
    def _dbg(m: str) -> None:
        if log: log(m)

    before = len(df)

    # 1. Drop rows where Symbol is blank, or the row looks like a header/schema
    #    placeholder (e.g. literal column names or type names like "string", "price")
    if "Symbol" in df.columns:
        df = df[df["Symbol"].astype(str).str.strip().str.len() > 0]
        junk_values = {
            "symbol", "contract", "name", "string", "integer", "number",
            "price", "pricechange", "float", "int", "text", "date", "time",
        }
        df = df[~df["Symbol"].astype(str).str.strip().str.lower().isin(junk_values)]
        if "Contract Name" in df.columns:
            df = df[~df["Contract Name"].astype(str).str.strip().str.lower().isin(junk_values)]
        if "Latest" in df.columns:
            df = df[~df["Latest"].astype(str).str.strip().str.lower().isin(junk_values)]

    # 2. Drop exact duplicates on Symbol (keep first occurrence)
    if "Symbol" in df.columns:
        df = df.drop_duplicates(subset=["Symbol"], keep="first")

    after_dedup = len(df)
    _dbg(f"  Deduplication: {before} → {after_dedup} rows")

    # 3. Convert fractional treasury prices in Latest and Change
    for col in ("Latest", "Change"):
        if col not in df.columns:
            continue
        converted = df[col].apply(_convert_treasury_price)
        changed = (converted != df[col]).sum()
        if changed:
            _dbg(f"  Converted {changed} values in '{col}' from fractional to decimal")
        df[col] = converted

    df = df.reset_index(drop=True)
    return df


def scrape_barchart_futures(url: str = BARCHART_URL, log: Any = None) -> pd.DataFrame:
    """
    Scrape the Barchart financials futures page for:
      Symbol, Contract Name, Latest, Change, Volume, Time

    Strategy (in order):
      1. Session-based proxima API call (fastest, no browser needed)
      2. Legacy marketdata.websol API call
      3. Playwright browser with network interception + cookie harvest
    """
    def _dbg(m: str) -> None:
        if log: log(m)

    _dbg("Building session (collecting cookies/XSRF)…")
    session = _barchart_session()

    _dbg("Strategy 1: Proxima internal API…")
    df = _barchart_via_proxima(session, log=log)
    if df is not None and not df.empty:
        return _postprocess_barchart(df, log=log)

    _dbg("Strategy 2: Legacy marketdata API…")
    df = _barchart_via_marketdata(session, log=log)
    if df is not None and not df.empty:
        return _postprocess_barchart(df, log=log)

    _dbg("Strategy 3: Playwright browser with network interception…")
    df = _barchart_via_playwright_intercept(url, log=log)
    if df is not None and not df.empty:
        return _postprocess_barchart(df, log=log)

    raise RuntimeError(
        "Could not extract futures data from Barchart.\n\n"
        "Check the log for details. All three strategies failed:\n"
        "  1. Proxima internal API – blocked or XSRF mismatch\n"
        "  2. Legacy marketdata API – key expired or blocked\n"
        "  3. Playwright interception – browser blocked or no JSON captured\n\n"
        "Make sure Playwright/Chromium is installed:\n"
        "  python -m pip install playwright\n"
        "  python -m playwright install chromium"
    )


def scrape_treasuries(url: str) -> pd.DataFrame:
    df = None

    try:
        df = fetch_with_requests(url)
    except Exception:
        df = None

    if df is None:
        df = extract_with_playwright(url)

    if df is None or df.empty:
        raise RuntimeError(
            "Could not extract Treasury table from page."
        )

    for col in df.columns:
        df[col] = df[col].astype(str).str.strip()

    missing = [c for c in TARGET_COLUMNS if c not in df.columns]
    if missing:
        raise RuntimeError(f"Missing expected columns after parsing: {missing}")

    return df[TARGET_COLUMNS].copy()

def macaulay_duration(
    settlement_date,
    maturity_date,
    coupon_rate,
    ask_yield,
    face_value=1000,
    frequency=2,):
    """
    Calculate Macaulay Duration for a bond.

    Inputs:
        settlement_date -> current/article date
        maturity_date   -> maturity date
        coupon_rate     -> annual coupon rate as decimal
        ask_yield       -> annual yield as decimal
        face_value      -> par value
        frequency       -> coupon frequency (2 for Treasuries)

    Returns:
        Macaulay Duration in years
    """

    settlement = parse_any_date(settlement_date)
    maturity = parse_any_date(maturity_date)

    # Total years to maturity
    years_to_maturity = (maturity - settlement).days / 365.25

    # Number of coupon periods
    n_periods = int(round(years_to_maturity * frequency))

    if n_periods <= 0:
        return 0.0

    # Periodic values
    y = ask_yield / frequency
    c = (coupon_rate * face_value) / frequency

    weighted_pv_sum = 0.0
    pv_total = 0.0

    for t in range(1, n_periods + 1):

        cash_flow = c

        # Final period includes principal repayment
        if t == n_periods:
            cash_flow += face_value

        pv = cash_flow / ((1 + y) ** t)

        weighted_pv_sum += t * pv
        pv_total += pv

    # Duration in periods
    duration_periods = weighted_pv_sum / pv_total

    # Convert periods -> years
    duration_years = duration_periods / frequency

    return duration_years

def modified_duration(
    settlement_date,
    maturity_date,
    coupon_rate,
    ask_yield,
    face_value=1000,
    frequency=2,
):
    """
    Calculate Modified Duration.

    Inputs:
        settlement_date -> current/article date
        maturity_date   -> maturity date
        coupon_rate     -> annual coupon rate as decimal
        ask_yield       -> annual yield as decimal
        face_value      -> bond par value
        frequency       -> coupon frequency (2 for Treasuries)

    Returns:
        Modified Duration in years
    """

    mac_dur = macaulay_duration(
        settlement_date=settlement_date,
        maturity_date=maturity_date,
        coupon_rate=coupon_rate,
        ask_yield=ask_yield,
        face_value=face_value,
        frequency=frequency,
    )

    y = ask_yield

    return mac_dur / (1 + y / frequency)
# ============================================================
# OUTPUT ENRICHMENT
# ============================================================

def add_payment_columns(
    df: pd.DataFrame,
    reference_date: date,
    delta: float = 0.01,
) -> pd.DataFrame:
    """Compute bond analytics columns.

    Parameters
    ----------
    df             : DataFrame from scrape_treasuries()
    reference_date : settlement / article date
    delta          : yield shift used for PV1/P1/PVUP/PVDN/EF_DURATION.
                     Expressed as a decimal fraction, e.g. 0.01 = 1%, 0.007 = 0.7%.
                     Default is 0.01 (1%).
    """
    df = df.copy()

    # Convert delta to percentage-point units (same scale as Asked Yield column)
    # Asked Yield is stored as e.g. 4.31 (meaning 4.31%), so delta_pp = 0.007*100 = 0.7
    delta_pp = delta * 100.0

    df["Delta"] = delta          # store the raw decimal value for reference
    df["Article Date"] = reference_date

    df["Previous Payment Date"] = df["Maturity"].apply(
        lambda x: closest_past_payment_date(x, reference_date)
    )

    df["Days Since Previous Payment"] = df["Maturity"].apply(
        lambda x: days_since_last_payment(x, reference_date)
    )

    df["Payments Until Maturity"] = df["Maturity"].apply(
        lambda x: count_payments_until_maturity(x, reference_date)
    )

    df["Coupon"] = pd.to_numeric(df["Coupon"], errors="coerce")
    df["Asked Yield"] = pd.to_numeric(df["Asked Yield"], errors="coerce")

    df["Coupon Payment"] = (df["Coupon"] / 2.0) * 10

    df["PV0"] = df.apply(
        lambda row: bond_present_value(
            ask_yield=row["Asked Yield"],
            payments_until_maturity=row["Payments Until Maturity"],
            coupon_payment=row["Coupon Payment"],
            face_value=1000,
        ),
        axis=1,
    )
    df["PV1"] = df.apply(
        lambda row: bond_future_value(
            ask_yield=row["Asked Yield"] + delta_pp,
            payments_until_maturity=row["Payments Until Maturity"],
            coupon_payment=row["Coupon Payment"],
            face_value=1000,
        ),
        axis=1,
    )
    df["P0"] = df["PV0"] * (1 + df["Asked Yield"] / 200) ** (df["Days Since Previous Payment"] / 182)
    df["P1"] = df["PV1"] * (1 + (df["Asked Yield"] + delta_pp) / 200) ** (df["Days Since Previous Payment"] / 182)
    df["Simulated Return"] = (df["P1"] - df["P0"] + df["Coupon"] * 10) / df["P0"]
    df["MACAULAY DURATION"] = df.apply(
        lambda row: macaulay_duration(
            settlement_date=reference_date,
            maturity_date=row["Maturity"],
            coupon_rate=row["Coupon"] / 100,
            ask_yield=row["Asked Yield"] / 100,
            face_value=1000,
            frequency=2,
        ),
        axis=1,
    )
    df["MODIFIED DURATION"] = df.apply(
        lambda row: modified_duration(
            settlement_date=reference_date,
            maturity_date=row["Maturity"],
            coupon_rate=row["Coupon"] / 100,
            ask_yield=row["Asked Yield"] / 100,
            face_value=1000,
            frequency=2,
        ),
        axis=1,
    )
    df["PVUP"] = df.apply(
        lambda row: bond_present_value(
            ask_yield=row["Asked Yield"] + delta_pp,
            payments_until_maturity=row["Payments Until Maturity"],
            coupon_payment=row["Coupon Payment"],
            face_value=1000,
        ),
        axis=1,
    )
    df["PVDN"] = df.apply(
        lambda row: bond_present_value(
            ask_yield=row["Asked Yield"] - delta_pp,
            payments_until_maturity=row["Payments Until Maturity"],
            coupon_payment=row["Coupon Payment"],
            face_value=1000,
        ),
        axis=1,
    )
    df["PUP"] = df["PVUP"] * (1 + (df["Asked Yield"] + delta_pp) / 200) ** (df["Days Since Previous Payment"] / 182)
    df["PDN"] = df["PVDN"] * (1 + (df["Asked Yield"] - delta_pp) / 200) ** (df["Days Since Previous Payment"] / 182)
    df["EF_DURATION"] = (df["PDN"] - df["PUP"]) / (2 * delta * df["P0"])
    return df


# ============================================================
# WSJ BOND REGRESSIONS
# ============================================================

def run_wsj_regressions(df: pd.DataFrame) -> pd.DataFrame:
    """
    Run 4 OLS regressions on WSJ Treasury data.
    Dependent variable : Asked Yield
    Independent vars   : Macaulay Duration, Modified Duration,
                         EF_DURATION, Simulated Return (one each)

    Returns a tidy summary DataFrame with one block of rows per model,
    suitable for writing directly to an Excel sheet.
    """
    try:
        import statsmodels.formula.api as smf
    except ImportError as e:
        raise ImportError(
            "statsmodels is required for regressions. "
            "Run: pip install statsmodels"
        ) from e

    dep_var = "Asked_Yield"   # statsmodels formula-safe name (no spaces)
    models = [
        ("Macaulay Duration",  "MACAULAY_DURATION"),
        ("Modified Duration",  "MODIFIED_DURATION"),
        ("Effective Duration", "EF_DURATION"),
        ("Simulated Return",   "Simulated_Return"),
    ]

    # Rename columns to formula-safe names for statsmodels
    rename = {
        "Asked Yield":       dep_var,
        "MACAULAY DURATION": "MACAULAY_DURATION",
        "MODIFIED DURATION": "MODIFIED_DURATION",
        "EF_DURATION":       "EF_DURATION",
        "Simulated Return":  "Simulated_Return",
    }
    reg_df = df.rename(columns=rename)

    all_cols = [dep_var] + [v for _, v in models]
    missing  = [c for c in all_cols if c not in reg_df.columns]
    if missing:
        raise ValueError(
            f"Regression columns not found: {missing}. "
            "Run scrape_treasuries + add_payment_columns first."
        )

    reg_df = reg_df[all_cols].dropna()

    summary_rows: list[dict] = []

    for label, indep_var in models:
        formula = f"{dep_var} ~ {indep_var}"
        fit     = smf.ols(formula, data=reg_df).fit()

        # Header row for this model
        summary_rows.append({
            "Model":       f"Ask Yield ~ {label}",
            "Metric":      "",
            "Coefficient": "",
            "Std Error":   "",
            "t-stat":      "",
            "p-value":     "",
            "R-squared":   "",
            "Adj R-sq":    "",
            "Observations":"",
        })

        # Stats row
        summary_rows.append({
            "Model":        "",
            "Metric":       "Model Stats",
            "Coefficient":  "",
            "Std Error":    "",
            "t-stat":       "",
            "p-value":      "",
            "R-squared":    round(float(fit.rsquared),     6),
            "Adj R-sq":     round(float(fit.rsquared_adj), 6),
            "Observations": int(fit.nobs),
        })

        # One row per coefficient (Intercept + slope)
        for param in fit.params.index:
            display = "Intercept" if param == "Intercept" else label
            summary_rows.append({
                "Model":        "",
                "Metric":       display,
                "Coefficient":  round(float(fit.params[param]),   6),
                "Std Error":    round(float(fit.bse[param]),      6),
                "t-stat":       round(float(fit.tvalues[param]),  4),
                "p-value":      round(float(fit.pvalues[param]),  6),
                "R-squared":    "",
                "Adj R-sq":     "",
                "Observations": "",
            })

        # Blank separator row
        summary_rows.append({k: "" for k in summary_rows[0]})

    return pd.DataFrame(summary_rows)


def build_regression_chart_image(df: pd.DataFrame) -> Optional[bytes]:
    """
    Build a 2×2 grid of scatter + regression line plots.
    One subplot per model: Asked Yield vs each of
      Macaulay Duration, Modified Duration, EF_DURATION, Simulated Return.
    Returns the chart as PNG bytes, or None if matplotlib is unavailable.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")   # non-interactive backend (no window)
        import matplotlib.pyplot as plt
        import matplotlib.gridspec as gridspec
        import numpy as np
        import io
    except ImportError:
        return None

    rename = {
        "Asked Yield":       "Asked Yield",
        "MACAULAY DURATION": "MACAULAY DURATION",
        "MODIFIED DURATION": "MODIFIED DURATION",
        "EF_DURATION":       "EF_DURATION",
        "Simulated Return":  "Simulated Return",
    }
    plot_df = df.rename(columns=rename)

    models = [
        ("Macaulay Duration",  "MACAULAY DURATION"),
        ("Modified Duration",  "MODIFIED DURATION"),
        ("Effective Duration", "EF_DURATION"),
        ("Simulated Return",   "Simulated Return"),
    ]

    fig = plt.figure(figsize=(14, 10))
    fig.suptitle("Ask Yield Regressions", fontsize=15, fontweight="bold", y=0.98)
    gs  = gridspec.GridSpec(2, 2, figure=fig, hspace=0.45, wspace=0.35)

    colors = ["#2196F3", "#4CAF50", "#FF9800", "#E91E63"]

    for idx, (label, col) in enumerate(models):
        ax = fig.add_subplot(gs[idx // 2, idx % 2])

        sub = plot_df[["Asked Yield", col]].dropna()
        if sub.empty or col not in sub.columns:
            ax.set_title(f"Ask Yield ~ {label}\n(no data)", fontsize=10)
            continue

        x = sub[col].values.astype(float)
        y = sub["Asked Yield"].values.astype(float)

        # Scatter
        ax.scatter(x, y, color=colors[idx], alpha=0.65, s=40, zorder=3,
                   edgecolors="white", linewidth=0.5)

        # Regression line
        if len(x) >= 2:
            m, b   = np.polyfit(x, y, 1)
            x_line = np.linspace(x.min(), x.max(), 200)
            ax.plot(x_line, m * x_line + b, color=colors[idx],
                    linewidth=2, zorder=4)

            # R²
            y_hat = m * x + b
            ss_res = np.sum((y - y_hat) ** 2)
            ss_tot = np.sum((y - y.mean()) ** 2)
            r2 = 1 - ss_res / ss_tot if ss_tot > 0 else float("nan")

            # Equation + R² annotation
            sign  = "+" if b >= 0 else "-"
            eq    = f"y = {m:.4f}x {sign} {abs(b):.4f}"
            r2txt = f"R² = {r2:.4f}"
            ax.text(0.05, 0.93, eq,    transform=ax.transAxes, fontsize=8,
                    color="black", verticalalignment="top")
            ax.text(0.05, 0.82, r2txt, transform=ax.transAxes, fontsize=9,
                    color=colors[idx], fontweight="bold", verticalalignment="top")

        ax.set_title(f"Ask Yield  ~  {label}", fontsize=10, fontweight="bold", pad=6)
        ax.set_xlabel(label, fontsize=8)
        ax.set_ylabel("Ask Yield (%)", fontsize=8)
        ax.tick_params(labelsize=7)
        ax.grid(True, linestyle="--", alpha=0.35, zorder=0)
        ax.set_axisbelow(True)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


# ============================================================
# MAIN
# ============================================================

def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Scrape WSJ Treasury or Barchart Futures quotes to Excel")
    parser.add_argument(
        "--source",
        choices=["wsj", "barchart"],
        default="wsj",
        help="Data source: 'wsj' (default) or 'barchart'",
    )
    parser.add_argument("--url", default=None, help="Override the default URL for the chosen source")
    parser.add_argument("--output", default=None, help="Output Excel file path (default depends on source)")
    parser.add_argument(
        "--reference-date",
        default=None,
        help="(WSJ only) Override the article date with YYYY-MM-DD",
    )
    parser.add_argument(
        "--delta",
        type=float,
        default=0.01,
        help=(
            "(WSJ only) Yield shift for PV1, P1, PVUP, PVDN, EF_DURATION. "
            "Expressed as a decimal fraction: 0.01 = 1%% (default), 0.007 = 0.7%%."
        ),
    )
    args = parser.parse_args(list(argv) if argv is not None else None)

    # ── Barchart path ──────────────────────────────────────────────────────────
    if args.source == "barchart":
        url = args.url or BARCHART_URL
        output_path = Path(args.output or "barchart_futures.xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        print(f"Scraping Barchart futures from {url} ...")
        df = scrape_barchart_futures(url, log=lambda m: print(m))

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Barchart Futures")

        print(f"Saved {len(df)} rows to {output_path.resolve()}")
        return 0

    # ── WSJ path ───────────────────────────────────────────────────────────────
    url = args.url or DEFAULT_URL
    output_path = Path(args.output or "wsj_treasuries.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    reference_date = get_reference_date(url, args.reference_date)
    df = scrape_treasuries(url)
    df = add_payment_columns(df, reference_date, delta=args.delta)

    delta_pct = args.delta * 100
    # Row layout:
    #   Row 1  → Date header (reference date or today)
    #   Row 2  → blank separator
    #   Rows 3–16 → formula legend (14 rows)
    #   Row 17+   → data header + rows
    DATE_ROW    = 1
    FORMULA_ROW = 3
    formula_rows = 16    # data starts after row 16 (0-indexed = startrow=16)
    sheet_name = "Treasuries"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=formula_rows)
        ws = writer.book[sheet_name]

        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import date as _date

        # ── Row 1: Date header ────────────────────────────────────────────────
        display_date = reference_date.strftime("%B %d, %Y")
        date_label   = f"Reference Date: {display_date}"
        ws["A1"] = date_label
        ws.merge_cells("A1:J1")   # merge across 10 columns so text is always fully visible
        ws["A1"].font      = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill      = PatternFill("solid", fgColor="1F4E79")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center",
                                       indent=1, wrap_text=False)
        ws.row_dimensions[1].height = 30   # taller row so 16pt text isn't clipped

        # ── Rows 3–16: Formula legend ─────────────────────────────────────────
        ws["A3"]  = "Treasury Formulas"
        ws["A3"].font = Font(bold=True)
        ws["A4"]  = "Coupon Payment";    ws["B4"]  = "Coupon/2*1000"
        ws["A5"]  = "PV0";               ws["B5"]  = "PV(Ask Yield/2, Number of Payments Until Maturity, Coupon Payment, 1000)"
        ws["A6"]  = "PV1";               ws["B6"]  = f"PV((Ask Yield+{delta_pct}%)/2, Number of Payments Until Maturity-2, Coupon Payment, 1000)"
        ws["A7"]  = "P0";                ws["B7"]  = "PV0 * (1+Asked Yield/2)^(Days Since Last Payment/182)"
        ws["A8"]  = "P1";                ws["B8"]  = f"PV1 * (1+(Asked Yield+{delta_pct}%)/2)^(Days Since Last Payment/182)"
        ws["A9"]  = "Simulated Return";  ws["B9"]  = "(P1- P0 + Coupon*10)/P0"
        ws["A10"] = "MACAULAY DURATION"; ws["B10"] = "DURATION(Current Date, Maturity Date, Coupon Rate, Ask Yield, 2)"
        ws["A11"] = "MODIFIED DURATION"; ws["B11"] = "MDURATION(Current Date, Maturity Date, Coupon Rate, Ask Yield, 2)"
        ws["A12"] = "PVUP";              ws["B12"] = f"PV(Ask Yield+{delta_pct}%/2, Number of Payments Until Maturity, Coupon Payment, 1000)"
        ws["A13"] = "PVDN";              ws["B13"] = f"PV(Ask Yield-{delta_pct}%/2, Number of Payments Until Maturity, Coupon Payment, 1000)"
        ws["A14"] = "PUP";               ws["B14"] = f"PVUP * (1+(Asked Yield+{delta_pct}%)/2)^(Days Since Last Payment/182)"
        ws["A15"] = "PDN";               ws["B15"] = f"PVDN * (1+(Asked Yield-{delta_pct}%)/2)^(Days Since Last Payment/182)"
        ws["A16"] = "EFDURATION";        ws["B16"] = f"(PDN- PUP)/(2*{args.delta}*P0)"

        # ── Sheet 2: Regressions ──────────────────────────────────────────────
        try:
            reg_summary = run_wsj_regressions(df)
            reg_summary.to_excel(writer, index=False, sheet_name="Regressions")
            reg_ws = writer.book["Regressions"]

            # Bold the model header rows (col A non-empty)
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill  = PatternFill("solid", fgColor="D9E1F2")
            bold_font    = Font(bold=True)
            for row in reg_ws.iter_rows():
                if row[0].value and str(row[0].value).startswith("Ask Yield"):
                    for cell in row:
                        cell.font = bold_font
                        cell.fill = header_fill
            # Auto-fit column widths
            for col in reg_ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                reg_ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
            print("  Regression sheet written.")

            # ── Sheet 3: Regression Charts ────────────────────────────────────
            chart_png = build_regression_chart_image(df)
            if chart_png is not None:
                from openpyxl.drawing.image import Image as XLImage
                import io
                chart_ws = writer.book.create_sheet("Regression Charts")
                chart_ws["A1"] = "Ask Yield Regression Plots"
                chart_ws["A1"].font = Font(bold=True, size=13)
                img = XLImage(io.BytesIO(chart_png))
                img.anchor = "A3"
                chart_ws.add_image(img)
                print("  Regression chart sheet written.")
        except Exception as reg_exc:
            print(f"  Warning: could not write regression sheet — {reg_exc}")

    print(f"Saved {len(df)} rows to {output_path.resolve()}")
    print(f"Reference date used: {reference_date.isoformat()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())